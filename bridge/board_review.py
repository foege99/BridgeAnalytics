import re

import pandas as pd
import numpy as np

from bridge.opening_bid import suggest_opening_for_row


# ---------------------------------------------------------------------------
# Direction helpers shared by board-layout functions
# ---------------------------------------------------------------------------

# Maps player-column name to compass direction code
_PLAYER_COL_TO_DIR = {'ns1': 'N', 'ns2': 'S', 'ew1': 'Ø', 'ew2': 'V'}

# Rotation tables: Per's original direction → display positions
# display positions: bottom (Per), top (partner), left (opponent), right (opponent)
_ROTATIONS = {
    'S': {'bottom': 'S', 'top': 'N', 'left': 'V', 'right': 'Ø'},
    'N': {'bottom': 'N', 'top': 'S', 'left': 'Ø', 'right': 'V'},
    'Ø': {'bottom': 'Ø', 'top': 'V', 'left': 'S', 'right': 'N'},
    'V': {'bottom': 'V', 'top': 'Ø', 'left': 'N', 'right': 'S'},
}

# Vulnerability → Danish display text
_VUL_DK = {
    '-':    'Ingen i zonen',
    'NS':   'NS i zonen',
    'ØV':   'ØV i zonen',
    'EW':   'ØV i zonen',
    'Alle': 'Alle i zonen',
}

# Target partner name for last-tournament traveller fallback logic
_PARTNER_NAME = "Henrik Friis"


def _hand_suit_lines(hand_str) -> list:
    """Convert dot-format hand string (S.H.D.C) to 4 suit lines with symbols."""
    if hand_str is None or (isinstance(hand_str, float) and pd.isna(hand_str)):
        parts = ['-', '-', '-', '-']
    else:
        parts = str(hand_str).split('.')
        while len(parts) < 4:
            parts.append('-')
    suits = ['♠', '♥', '♦', '♣']
    return [f"{s} {p}" for s, p in zip(suits, parts)]


def _both_names_in_df(df_subset: pd.DataFrame, name1: str, name2: str) -> bool:
    """Return True if both *name1* and *name2* appear (exact strip match) in any player column."""
    n1 = name1.strip() if name1 else ''
    n2 = name2.strip() if name2 else ''
    if not n1 or not n2:
        return False
    found1 = found2 = False
    for col in ('ns1', 'ns2', 'ew1', 'ew2'):
        if col not in df_subset.columns:
            continue
        for val in df_subset[col]:
            if val is None or (isinstance(val, float) and pd.isna(val)):
                continue
            s = str(val).strip()
            if s == n1:
                found1 = True
            if s == n2:
                found2 = True
            if found1 and found2:
                return True
    return False


def write_board1_layout_sheet(
    writer,
    df: pd.DataFrame,
    per_name: str,
    board_no: int = 1,
    sheet_name: str | None = None,
) -> None:
    """
    Write a single board layout sheet to *writer* (an open pd.ExcelWriter).

    Shows *board_no* from the latest tournament date present in *df* in a
    classic bridge table layout, rotated so that *per_name* is always at the
    bottom.

    Defaults preserve legacy behaviour:
    - board_no=1
    - sheet_name='Board1_LastTournament'

    If the required data is not available the sheet is still created with a
    descriptive message instead of raising an exception.
    """
    try:
        from openpyxl.styles import Font
    except ImportError:
        Font = None  # graceful degradation – no bold formatting

    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
        _RED_FONT = InlineFont(color='FF0000')
        _rich_text_available = True
    except ImportError:
        _rich_text_available = False

    try:
        target_board = int(board_no)
    except (TypeError, ValueError):
        target_board = 1

    wb = writer.book
    if sheet_name is None:
        sheet_name = (
            "Board1_LastTournament"
            if target_board == 1
            else f"Board{target_board}_LastTournament"
        )
    ws = wb.create_sheet(sheet_name)

    def _write_msg(msg: str) -> None:
        ws.cell(row=1, column=1, value=msg)

    def _bold(cell):
        if Font is not None:
            cell.font = Font(bold=True)
        return cell

    def _write_with_red_suits(cell, text) -> None:
        """Write *text* to *cell*; colour any ♥ or ♦ characters red using rich text."""
        if text is None:
            cell.value = text
            return
        text_str = str(text)
        if not _rich_text_available or ('♥' not in text_str and '♦' not in text_str):
            cell.value = text_str
            return
        parts = []
        i = 0
        while i < len(text_str):
            if text_str[i] in ('♥', '♦'):
                parts.append(TextBlock(_RED_FONT, text_str[i]))
                i += 1
            else:
                j = i + 1
                while j < len(text_str) and text_str[j] not in ('♥', '♦'):
                    j += 1
                parts.append(text_str[i:j])
                i = j
        cell.value = CellRichText(parts)

    def _write_suit_line(cell, line: str) -> None:
        """Write a suit line; colour ♥ and ♦ symbols red when rich text is available."""
        if _rich_text_available and line and line[0] in ('♥', '♦'):
            cell.value = CellRichText([TextBlock(_RED_FONT, line[0]), line[1:]])
        else:
            cell.value = line

    # ------------------------------------------------------------------
    # 1. Validate input
    # ------------------------------------------------------------------
    if df is None or df.empty:
        _write_msg("Ingen data tilgængelig.")
        return

    if 'board_no' not in df.columns:
        _write_msg("Kolonnen 'board_no' mangler i datasættet.")
        return

    if 'tournament_date' not in df.columns:
        _write_msg("Kolonnen 'tournament_date' mangler i datasættet.")
        return

    # ------------------------------------------------------------------
    # 2. Latest tournament date → target board
    # ------------------------------------------------------------------
    latest_date = df['tournament_date'].max()
    df_latest = df[df['tournament_date'] == latest_date]
    df_board = df_latest[df_latest['board_no'] == target_board]

    if df_board.empty:
        _write_msg(
            f"Spil {target_board} ikke fundet for seneste turnering ({latest_date})."
        )
        return

    # ------------------------------------------------------------------
    # 3. Find the row where Per appears
    # ------------------------------------------------------------------
    per_row = None
    per_dir = None

    for _, row in df_board.iterrows():
        for col, dir_code in _PLAYER_COL_TO_DIR.items():
            if row.get(col) == per_name:
                per_row = row
                per_dir = dir_code
                break
        if per_row is not None:
            break

    if per_row is None:
        _write_msg(
            f"{per_name} ikke fundet i Spil {target_board} for seneste turnering ({latest_date})."
        )
        return

    # ------------------------------------------------------------------
    # 4. Build display data
    # ------------------------------------------------------------------
    rot = _ROTATIONS[per_dir]

    dir_to_player = {
        'N': per_row.get('ns1', ''),
        'S': per_row.get('ns2', ''),
        'Ø': per_row.get('ew1', ''),
        'V': per_row.get('ew2', ''),
    }
    dir_to_hand = {
        'N': per_row.get('N_hand'),
        'S': per_row.get('S_hand'),
        'Ø': per_row.get('Ø_hand'),
        'V': per_row.get('V_hand'),
    }

    # HCP lookup per direction
    _hcp_col_map = {'N': 'N_HCP', 'S': 'S_HCP', 'Ø': 'Ø_HCP', 'V': 'V_HCP'}

    def _get_hcp(dir_code: str):
        """Return HCP for *dir_code* as int, or '?' if unavailable.

        Priority: dd_{dir}_HCP (when dd_valid) → per-direction HCP column →
        computed from dot-format hand string.
        """
        # 1) DD HCP when available
        if per_row.get('dd_valid'):
            dd_col = f'dd_{dir_code}_HCP'
            if dd_col in per_row.index:
                val = per_row.get(dd_col)
                if val is not None and not (isinstance(val, float) and pd.isna(val)):
                    return int(val)
        # 2) Regular per-direction HCP column
        col = _hcp_col_map.get(dir_code, '')
        if col and col in per_row.index:
            val = per_row.get(col)
            if val is not None and not (isinstance(val, float) and pd.isna(val)):
                return int(val)
        # 3) Fallback: compute from hand string
        hand = dir_to_hand.get(dir_code)
        if hand is not None and not (isinstance(hand, float) and pd.isna(hand)):
            try:
                from bridge.hand_eval import parse_hand, hcp as _calc_hcp
                return _calc_hcp(parse_hand(str(hand)))
            except Exception:
                pass
        return '?'

    def _player_label(dir_code: str) -> str:
        name = dir_to_player.get(dir_code, '')
        return f"{dir_code}: {name}  {_get_hcp(dir_code)} HCP"

    # Graceful field lookup (tries multiple candidate column names)
    def _get_field(row, *candidates):
        """Return the first non-null value found in *row* for any of *candidates*.

        Parameters
        ----------
        row : pandas Series
            The data row to inspect.
        *candidates : str
            Column names to try in order.

        Returns
        -------
        The first non-null field value, or None if all candidates are missing/null.
        """
        for c in candidates:
            if c in row.index:
                val = row.get(c)
                if val is not None and not (isinstance(val, float) and pd.isna(val)):
                    return val
        return None

    top_dir = rot['top']
    left_dir = rot['left']
    right_dir = rot['right']
    bottom_dir = rot['bottom']  # always Per

    # ------------------------------------------------------------------
    # 5. Write to sheet
    #
    # Layout (columns A=1, B=2, C=3, E=5):
    #   Row 1 : title (B)  |  Turnering (E)
    #   Row 2 : Board (A)  |  top player name (B)  |  Kontrakt (C)
    #   Row 3 : Dealer (A)  |  top hand suit 1 (B)  |  Spilfører (C)
    #   Row 4 : NS HCP (A)  |  top hand suit 2 (B)  |  Udspil (C)  |  Zone (E)
    #   Row 5 : ØV HCP (A)  |  top hand suit 3 (B)  |  Resultat (C)
    #   Row 6 : top hand suit 4 (B)
    #   Row 8 : left name (A)  |  right name (C)
    #   Rows 9-12 : left suits (A)  |  right suits (C)
    #   Row 14 : bottom player name (col B)
    #   Rows 15-18 : bottom hand suits (col B)
    # ------------------------------------------------------------------
    section_val = per_row.get('row', per_row.get('section', ''))
    ws.cell(row=1, column=2,
            value=f"Spil {target_board} – {latest_date} (sektion {section_val})")
    _bold(ws.cell(row=1, column=2))

    # --- Header block ---
    dealer_val = _get_field(per_row, 'dealer', 'Dealer')
    zone_val = _get_field(per_row, 'vul', 'vulnerability', 'zone', 'Zone')
    zone_display = _VUL_DK.get(zone_val, zone_val) if zone_val is not None else '(ukendt)'
    # ws.cell(row=1, column=5, value=f"Turnering: {latest_date}")
    # ws.cell(row=2, column=1, value="Spil: 1")
    ws.cell(row=3, column=1,
            value=f"Giver: {dealer_val if dealer_val is not None else '(ukendt)'}")
    ws.cell(row=6, column=1, value=f"Zone: {zone_display}")

    # --- Top hand ---
    _bold(ws.cell(row=2, column=2, value=_player_label(top_dir)))
    for i, line in enumerate(_hand_suit_lines(dir_to_hand.get(top_dir))):
        _write_suit_line(ws.cell(row=3 + i, column=2), line)

    # --- Left hand ---
    _bold(ws.cell(row=8, column=1, value=_player_label(left_dir)))
    for i, line in enumerate(_hand_suit_lines(dir_to_hand.get(left_dir))):
        _write_suit_line(ws.cell(row=9 + i, column=1), line)

    # --- Right hand ---
    _bold(ws.cell(row=8, column=3, value=_player_label(right_dir)))
    for i, line in enumerate(_hand_suit_lines(dir_to_hand.get(right_dir))):
        _write_suit_line(ws.cell(row=9 + i, column=3), line)

    # --- Bottom hand (Per) ---
    _bold(ws.cell(row=14, column=2, value=_player_label(bottom_dir)))
    for i, line in enumerate(_hand_suit_lines(dir_to_hand.get(bottom_dir))):
        _write_suit_line(ws.cell(row=15 + i, column=2), line)

    # --- Info block (A2-A5, C2-C5) ---
    contract_val = _get_field(per_row, 'contract')
    decl_val = _get_field(per_row, 'decl')
    lead_val = _get_field(per_row, 'lead')
    tricks_val = _get_field(per_row, 'tricks')

    def _side_hcp_total(col_name: str, dir1: str, dir2: str):
        """Return combined HCP for a side.

        Priority: dd HCP sum (when dd_valid) → combined column → sum per-direction HCP.
        """
        if per_row.get('dd_valid'):
            dd1 = per_row.get(f'dd_{dir1}_HCP')
            dd2 = per_row.get(f'dd_{dir2}_HCP')
            if (dd1 is not None and not (isinstance(dd1, float) and pd.isna(dd1)) and
                    dd2 is not None and not (isinstance(dd2, float) and pd.isna(dd2))):
                try:
                    return int(dd1) + int(dd2)
                except (ValueError, TypeError):
                    pass
        val = _get_field(per_row, col_name)
        if val is not None:
            return val
        h1, h2 = _get_hcp(dir1), _get_hcp(dir2)
        if h1 != '?' and h2 != '?':
            return int(h1) + int(h2)
        return None

    # NS/ØV HCP totals: prefer combined columns, else sum per-hand
    ns_hcp = _side_hcp_total('NS_HCP', 'N', 'S')
    ov_hcp = _side_hcp_total('ØV_HCP', 'Ø', 'V')

    ws.cell(row=4, column=1, value=f"NS HCP: {ns_hcp if ns_hcp is not None else '(ukendt)'}")
    ws.cell(row=5, column=1, value=f"ØV HCP: {ov_hcp if ov_hcp is not None else '(ukendt)'}")

    _write_with_red_suits(ws.cell(row=2, column=3),
                          f"Kontrakt: {contract_val if contract_val is not None else '(ukendt)'}")
    ws.cell(row=3, column=3,
            value=f"Spilfører: {decl_val if decl_val is not None else '(ukendt)'}")
    _write_with_red_suits(ws.cell(row=4, column=3),
                          f"Udspil: {lead_val if lead_val is not None else '(ukendt)'}")
    ws.cell(row=5, column=3,
            value=f"Resultat: {tricks_val if tricks_val is not None else '(ukendt)'}")

    # Par/DD-best contract note moved to O9 (requested placement)
    par_score = _get_field(per_row, 'par_score')
    par_contract = _get_field(per_row, 'par_contract')
    par_side = _get_field(per_row, 'par_side')
    if par_score is not None and par_contract is not None:
        par_text = f"Par: {par_score} {par_contract}"
        if par_side:
            par_text += f" {par_side}"
    else:
        par_text = None
    ws.cell(row=9, column=15,
            value=par_text if par_text is not None else "Par: (ukendt)")

    # --- Column widths ---
    # Reserve A:D for the bidding sequence scaffold.
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['O'].width = 32

       # ------------------------------------------------------------------
    # 6. Traveller table (ALL results for same tournament_date + row + board_no)
    # ------------------------------------------------------------------
    try:
        from openpyxl.styles import PatternFill, Border, Side, Alignment
        _styles_available = True
    except ImportError:
        _styles_available = False

    _GRAY_FILL = 'D9D9D9'  # shared fill color for header cells
    _WHITE_FILL = 'FFFFFF'
    _HILITE_YELLOW = 'FFF2CC'
    _ZEBRA_FILL = 'F7F7F7'

    # Bidding scaffold (visual placeholder, actual auction content comes later)
    _BID_HDR_ROW = 20
    _BID_DATA_START_ROW = 21
    _BID_DATA_ROWS = 12
    _BID_LOG_START_ROW = _BID_DATA_START_ROW + _BID_DATA_ROWS + 3  # ~15 rows below header
    _BID_START_COL = 1  # A
    _BID_HEADERS = ['S', 'V', 'N', 'Ø']
    _BID_LIGHT_GREEN = 'EAF2E3'
    _BID_LIGHT_PINK = 'F4E4E8'

    zone_norm = str(zone_val).strip().upper() if zone_val is not None else ''
    ns_vul = zone_norm in ('NS', 'ALLE', 'ALL')
    ov_vul = zone_norm in ('ØV', 'EW', 'ALLE', 'ALL')
    seat_side = {'S': 'NS', 'N': 'NS', 'V': 'ØV', 'Ø': 'ØV'}

    thin_bid = Side(style='thin') if _styles_available else None

    # Header row A20:D20 (bold + zone-aware color)
    for i, seat in enumerate(_BID_HEADERS):
        col_num = _BID_START_COL + i
        hdr_cell = ws.cell(row=_BID_HDR_ROW, column=col_num, value=seat)
        if Font is not None:
            hdr_cell.font = Font(bold=True)
        if _styles_available:
            hdr_cell.alignment = Alignment(horizontal='center', vertical='center')
            if thin_bid:
                hdr_cell.border = Border(left=thin_bid, right=thin_bid, top=thin_bid, bottom=thin_bid)
            is_vul = (
                (seat_side.get(seat) == 'NS' and ns_vul)
                or (seat_side.get(seat) == 'ØV' and ov_vul)
            )
            hdr_fill = _BID_LIGHT_PINK if is_vul else _BID_LIGHT_GREEN
            hdr_cell.fill = PatternFill(fill_type='solid', fgColor=hdr_fill)

    # Dealer opening suggestion from YAML profile logic (MVP: only first call).
    opening_guess = suggest_opening_for_row(per_row)
    opening_dealer = opening_guess.get('dealer')
    opening_display_bid = opening_guess.get('display_bid')
    opening_log_lines = opening_guess.get('log_lines') if isinstance(opening_guess, dict) else []
    bid_col_by_seat = {'S': 1, 'V': 2, 'N': 3, 'Ø': 4}
    if opening_dealer in bid_col_by_seat and opening_display_bid:
        ws.cell(
            row=_BID_DATA_START_ROW,
            column=bid_col_by_seat[opening_dealer],
            value=opening_display_bid,
        )

    # Opening decision log block shown below the bidding table.
    if isinstance(opening_log_lines, list) and opening_log_lines:
        ws.merge_cells(
            start_row=_BID_LOG_START_ROW,
            start_column=1,
            end_row=_BID_LOG_START_ROW,
            end_column=4,
        )
        log_hdr = ws.cell(row=_BID_LOG_START_ROW, column=1, value='Åbningslog')
        if Font is not None:
            log_hdr.font = Font(bold=True)
        if _styles_available:
            log_hdr.alignment = Alignment(horizontal='left', vertical='center')
            log_hdr.fill = PatternFill(fill_type='solid', fgColor=_BID_LIGHT_GREEN)
            if thin_bid:
                log_hdr.border = Border(left=thin_bid, right=thin_bid, top=thin_bid, bottom=thin_bid)

        for idx, line in enumerate(opening_log_lines[:10], start=1):
            r_log = _BID_LOG_START_ROW + idx
            ws.merge_cells(start_row=r_log, start_column=1, end_row=r_log, end_column=4)
            log_cell = ws.cell(row=r_log, column=1)
            _write_with_red_suits(log_cell, str(line))
            if _styles_available:
                log_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                if thin_bid:
                    log_cell.border = Border(left=thin_bid, right=thin_bid, top=thin_bid, bottom=thin_bid)

    # Body rows A21:D32 (light green background).
    # If bid text is present later, ensure hearts/diamonds render red.
    for r in range(_BID_DATA_START_ROW, _BID_DATA_START_ROW + _BID_DATA_ROWS):
        for c in range(_BID_START_COL, _BID_START_COL + len(_BID_HEADERS)):
            bid_cell = ws.cell(row=r, column=c)
            if _styles_available:
                bid_cell.alignment = Alignment(horizontal='center', vertical='center')
                if thin_bid:
                    bid_cell.border = Border(left=thin_bid, right=thin_bid, top=thin_bid, bottom=thin_bid)
                bid_cell.fill = PatternFill(fill_type='solid', fgColor=_BID_LIGHT_GREEN)
            if bid_cell.value is not None and str(bid_cell.value).strip():
                _write_with_red_suits(bid_cell, bid_cell.value)

    # Traveller headers (bridge.dk style)
    _TRAVELLER_HEADERS = [
        'NS', 'ØV', 'Kontrakt', 'Spilfører', 'Udspil', 'Stik',
        'Score NS', 'Score ØV', 'Point NS', 'Point ØV', 'Pct NS', 'Pct ØV',
        'Pct Defense', 'Pct Decl',
        'Lead type',
    ]
    _TRAV_START_COL = 5  # E
    _TRAV_HEADER_ROW = 1
    _TRAV_DATA_START_ROW = 2

    def _apply_header_style(cell) -> None:
        if not _styles_available:
            return
        cell.font = Font(bold=True) if Font is not None else cell.font
        cell.fill = PatternFill(fill_type='solid', fgColor=_GRAY_FILL)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        thin = Side(style='thin')
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _apply_data_style(cell, align: str = 'center', fill_color: str | None = None) -> None:
        if not _styles_available:
            return
        thin = Side(style='thin')
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=False)
        if Font is not None:
            cell.font = Font(color='000000')
        cell.fill = PatternFill(fill_type='solid', fgColor=fill_color if fill_color else _WHITE_FILL)

    # Filter to all results for this board/date/section
    cur_date = per_row.get('tournament_date')
    cur_board = per_row.get('board_no')
    cur_section = per_row.get('row', per_row.get('section'))

    # ------------------------------------------------------------------
    # Fallback tournament date: if both target names are not present for
    # cur_date + cur_section, use the most recent earlier date where both
    # participated (applies only to df_trav – the traveller table).
    # ------------------------------------------------------------------
    _trav_date = cur_date
    _fallback_note = None
    if cur_date is not None:
        _section_col = (
            'row' if 'row' in df.columns
            else 'section' if 'section' in df.columns
            else None
        )
        if _section_col is not None and cur_section is not None:
            _df_check = df[
                (df['tournament_date'] == cur_date) &
                (df[_section_col] == cur_section)
            ]
        else:
            _df_check = df[df['tournament_date'] == cur_date]

        if not _both_names_in_df(_df_check, per_name, _PARTNER_NAME):
            _df_sec = (
                df[df[_section_col] == cur_section]
                if _section_col is not None and cur_section is not None
                else df
            )
            _other_dates = sorted(
                (d for d in _df_sec['tournament_date'].unique() if d != cur_date),
                reverse=True,
            )
            for _d in _other_dates:
                _df_d = _df_sec[_df_sec['tournament_date'] == _d]
                if _both_names_in_df(_df_d, per_name, _PARTNER_NAME):
                    _fallback_note = (
                        f"OBS: Seneste turnering ({cur_date}) inkluderede ikke begge "
                        f"spillere. Viser sidst spillede turnering fra {_d}."
                    )
                    _trav_date = _d
                    break

    df_trav = df.copy()
    if _trav_date is not None:
        df_trav = df_trav[df_trav['tournament_date'] == _trav_date]
    if cur_board is not None and 'board_no' in df_trav.columns:
        df_trav = df_trav[df_trav['board_no'] == cur_board]
    if cur_section is not None and 'row' in df_trav.columns:
        df_trav = df_trav[df_trav['row'] == cur_section]

    # If we still ended with empty (should not happen), fall back to df_board
    if df_trav.empty:
        df_trav = df_board.copy()

    # Write fallback note to sheet if applicable (row 7, col B)
    if _fallback_note is not None:
        _note_cell = ws.cell(row=7, column=2, value=_fallback_note)
        if Font is not None:
            _note_cell.font = Font(italic=True)

    # Build traveller rows
    def _pair_text(r, side: str) -> str:
        if side == 'NS':
            val = _get_field(r, 'ns_pair')
            if val:
                return str(val)
            a = _get_field(r, 'ns1') or ''
            b = _get_field(r, 'ns2') or ''
            return " - ".join([p for p in [a, b] if p])
        else:
            val = _get_field(r, 'ew_pair')
            if val:
                return str(val)
            a = _get_field(r, 'ew1') or ''
            b = _get_field(r, 'ew2') or ''
            return " - ".join([p for p in [a, b] if p])

    def _lead_type_text(r) -> str:
        return _lead_type_text_from_row(r)

    def _pct_or_unknown(val):
        """Return pct value or 'ukendt' when unavailable."""
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return 'ukendt'
        return val

    def _to_number_or_none(val):
        """Try parse a numeric value; return None when unavailable/non-numeric."""
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return None
        if isinstance(val, (int, float, np.integer, np.floating)):
            if pd.isna(val):
                return None
            return float(val)
        s = str(val).strip().replace("\xa0", "").replace(" ", "")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None

    def _as_excel_number(num):
        """Return int when whole-number, else float (or None)."""
        if num is None:
            return None
        if float(num).is_integer():
            return int(num)
        return float(num)

    def _mirrored_scores(r):
        """Return (score_ns, score_ov) with one-sided score mirrored when possible."""
        score_ns_raw = _get_field(r, 'score_NS', 'score_ns', 'NS_score', 'score')
        score_ov_raw = _get_field(r, 'score_ØV', 'score_ew', 'ØV_score', 'score_EW')

        score_ns_num = _to_number_or_none(score_ns_raw)
        score_ov_num = _to_number_or_none(score_ov_raw)

        if score_ns_num is None and score_ov_num is not None:
            score_ns_num = -score_ov_num
        if score_ov_num is None and score_ns_num is not None:
            score_ov_num = -score_ns_num

        # Prefer numeric mirrored values when parseable; else keep original text values.
        score_ns_out = _as_excel_number(score_ns_num) if score_ns_num is not None else score_ns_raw
        score_ov_out = _as_excel_number(score_ov_num) if score_ov_num is not None else score_ov_raw
        return score_ns_out, score_ov_out

    # Write header
    for i, header in enumerate(_TRAVELLER_HEADERS):
        cell = ws.cell(row=_TRAV_HEADER_ROW, column=_TRAV_START_COL + i, value=header)
        _apply_header_style(cell)

    # Determine highlight row: any row containing both names in either NS or ØV text
    def _is_target_pair(ns_txt: str, ew_txt: str) -> bool:
        blob = f"{ns_txt} {ew_txt}".lower()
        return ("henrik friis" in blob) and ("per føge jensen" in blob)

    target_pair_lead_value = None
    target_pair_contract_color = None
    target_pair_dirs: set[str] = set()

    # Write data rows
    for ridx, (_, r) in enumerate(df_trav.iterrows(), start=0):
        out_row = _TRAV_DATA_START_ROW + ridx

        ns_txt = _pair_text(r, 'NS')
        ew_txt = _pair_text(r, 'EW')
        score_ns_val, score_ov_val = _mirrored_scores(r)

        values_and_align = [
            (ns_txt, 'left'),
            (ew_txt, 'left'),
            (_get_field(r, 'contract'), 'center'),
            (_normalize_compass(_get_field(r, 'decl')), 'center'),
            (_get_field(r, 'lead'), 'center'),
            (_get_field(r, 'tricks'), 'right'),
            (score_ns_val, 'right'),
            (score_ov_val, 'right'),
            (_get_field(r, 'point_NS', 'mp_NS', 'imp_NS', 'mp_ns', 'imp_ns'), 'right'),
            (_get_field(r, 'point_ØV', 'mp_ØV', 'imp_ØV', 'mp_ew', 'imp_ew'), 'right'),
            (_get_field(r, 'pct_NS', 'pct_ns'), 'right'),
            (_get_field(r, 'pct_ØV', 'pct_EW', 'pct_ew'), 'right'),
            (_pct_or_unknown(_pct_defense_from_row(r)), 'right'),
            (_pct_or_unknown(_pct_decl_from_row(r)), 'right'),
            (_lead_type_text(r), 'left'),
        ]

        # Zebra + highlight
        fill = None
        if _is_target_pair(ns_txt, ew_txt):
            fill = _HILITE_YELLOW
            target_pair_lead_value = _get_field(r, 'lead')
            target_pair_contract_color = _contract_pool_to_color_symbol(
                _contract_pool_from_row(r)
            )
            ns_blob = str(ns_txt).lower()
            ew_blob = str(ew_txt).lower()
            if ("henrik friis" in ns_blob) and ("per føge jensen" in ns_blob):
                target_pair_dirs = {'N', 'S'}
            elif ("henrik friis" in ew_blob) and ("per føge jensen" in ew_blob):
                target_pair_dirs = {'Ø', 'V'}
        elif (ridx % 2) == 1:
            fill = _ZEBRA_FILL

        for cidx, (val, align) in enumerate(values_and_align):
            cell = ws.cell(row=out_row, column=_TRAV_START_COL + cidx)
            # Contract (index 2) and lead/Udspil (index 4) may contain ♥/♦
            if cidx in (2, 4):
                _write_with_red_suits(cell, val)
            else:
                cell.value = val
            _apply_data_style(cell, align=align, fill_color=fill)

    # Column widths for traveller columns (tuned for reliable header readability)
    _TRAV_COL_WIDTHS = [28, 30, 12, 9, 10, 6, 12, 12, 10, 10, 8, 8, 11, 11, 28]
    _col_letters = 'EFGHIJKLMNOPQRS'
    for letter, width in zip(_col_letters, _TRAV_COL_WIDTHS):
        ws.column_dimensions[letter].width = width
    ws.row_dimensions[_TRAV_HEADER_ROW].height = 20

    # Compute DD start row: traveller header + traveller rows + 2 blank rows
    _DD_START_ROW = _TRAV_HEADER_ROW + 1 + len(df_trav) + 2
    # _DD_START_COL = 5
    # E    # ------------------------------------------------------------------
    # 7. Double Dummy table  E6:K10
    # ------------------------------------------------------------------
    _DD_START_COL = 5  # E

    dd_valid = per_row.get('dd_valid')
    if dd_valid is None or (isinstance(dd_valid, float) and pd.isna(dd_valid)):
        dd_valid = False

    if not dd_valid:
        na_cell = ws.cell(row=_DD_START_ROW, column=_DD_START_COL,
                          value='Double Dummy: ikke tilgængelig')
        if Font is not None:
            na_cell.font = Font(italic=True)
    else:
        if not target_pair_dirs:
            partner_dir = {'N': 'S', 'S': 'N', 'Ø': 'V', 'V': 'Ø'}.get(per_dir)
            if partner_dir is not None:
                target_pair_dirs = {per_dir, partner_dir}

        # Title
        title_cell = ws.cell(row=_DD_START_ROW, column=_DD_START_COL,
                             value='Double Dummy')

        _DD_STRAIN_HEADERS = ['NT', '♠', '♥', '♦', '♣', 'HP']
        _DD_DIRS_ORDER = ['N', 'S', 'Ø', 'V']
        _DD_STRAIN_KEYS = ['NT', 'S', 'H', 'D', 'C']  # column keys for dd_{dir}_{key}

        thin = Side(style='thin') if _styles_available else None

        def _dd_cell_style(cell, is_header: bool = False,
                           is_ns: bool = False,
                           fill_color: str | None = None) -> None:
            if not _styles_available:
                return
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if thin:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if fill_color is not None:
                cell.fill = PatternFill(fill_type='solid', fgColor=fill_color)
                if is_header and Font is not None:
                    cell.font = Font(bold=True)
            elif is_header and Font is not None:
                cell.fill = PatternFill(fill_type='solid', fgColor=_GRAY_FILL)
                cell.font = Font(bold=True)
            elif is_ns:
                cell.fill = PatternFill(fill_type='solid', fgColor='EBF1DE')

        # Title cell should look like DD table headers.
        _dd_cell_style(title_cell, is_header=True)

        # Header row: H6:M6
        for j, strain_hdr in enumerate(_DD_STRAIN_HEADERS):
            c = ws.cell(row=_DD_START_ROW, column=_DD_START_COL + 1 + j)
            _write_with_red_suits(c, strain_hdr)
            _dd_cell_style(c, is_header=True)

        # Data rows: G7:M10
        for i, dir_code in enumerate(_DD_DIRS_ORDER):
            row_num = _DD_START_ROW + 1 + i
            is_ns = dir_code in ('N', 'S')
            row_fill = _HILITE_YELLOW if dir_code in target_pair_dirs else None

            # Row label
            label_cell = ws.cell(row=row_num, column=_DD_START_COL,
                                 value=dir_code)
            _dd_cell_style(label_cell, is_header=True)

            # Strain values
            for j, strain_key in enumerate(_DD_STRAIN_KEYS):
                col_name = f'dd_{dir_code}_{strain_key}'
                val = per_row.get(col_name)
                if val is not None and not (isinstance(val, float) and pd.isna(val)):
                    try:
                        val = int(val)
                    except (ValueError, TypeError):
                        pass
                else:
                    val = None
                dc = ws.cell(row=row_num, column=_DD_START_COL + 1 + j, value=val)
                _dd_cell_style(dc, is_ns=is_ns, fill_color=row_fill)

            # HCP column (M = index 5)
            hcp_col = f'dd_{dir_code}_HCP'
            hcp_val = per_row.get(hcp_col)
            if hcp_val is not None and not (isinstance(hcp_val, float) and pd.isna(hcp_val)):
                try:
                    hcp_val = int(hcp_val)
                except (ValueError, TypeError):
                    pass
            else:
                hcp_val = None
            hc = ws.cell(row=row_num, column=_DD_START_COL + 6, value=hcp_val)
            _dd_cell_style(hc, is_ns=is_ns, fill_color=row_fill)

    # ------------------------------------------------------------------
    # 8. Lead-effekt (pooled A+B+C for current board + tournament)
    # ------------------------------------------------------------------
    _dd_end_row = _DD_START_ROW + 4 if dd_valid else _DD_START_ROW
    _LEAD_START_ROW = _dd_end_row + 2
    _LEAD_START_COL = 5  # E

    df_lead_pool = df.copy()
    if cur_date is not None:
        df_lead_pool = df_lead_pool[df_lead_pool['tournament_date'] == cur_date]
    if cur_board is not None and 'board_no' in df_lead_pool.columns:
        df_lead_pool = df_lead_pool[df_lead_pool['board_no'] == cur_board]

    _row_col = (
        'row' if 'row' in df_lead_pool.columns
        else 'section' if 'section' in df_lead_pool.columns
        else None
    )

    rows_present_txt = 'A+B+C'
    if _row_col is not None:
        df_lead_pool = df_lead_pool.copy()
        df_lead_pool['_row_code'] = df_lead_pool[_row_col].apply(_normalize_row_code)
        df_lead_pool = df_lead_pool[df_lead_pool['_row_code'].isin(['A', 'B', 'C'])]
        _rows_present = sorted(df_lead_pool['_row_code'].dropna().unique().tolist())
        if _rows_present:
            rows_present_txt = '+'.join(_rows_present)

    total_leads = len(df_lead_pool)
    df_lead_valid = _filter_valid_leads(df_lead_pool)
    valid_leads = len(df_lead_valid)
    lead_summary_rows = _make_lead_effect_summary_rows(
        df_lead_valid,
        contract_top_n=5,
        include_decl_hand=True,
    )

    show_contract_pool = any('contract_pool' in r for r in lead_summary_rows)
    show_decl_hand = any('decl_hand' in r for r in lead_summary_rows)

    title_cell = ws.cell(
        row=_LEAD_START_ROW,
        column=_LEAD_START_COL,
        value=f"Lead-effekt (pooled {rows_present_txt}, top-5 kontraktpool)",
    )
    _bold(title_cell)

    ws.cell(
        row=_LEAD_START_ROW + 1,
        column=_LEAD_START_COL,
        value=f"Gyldige udspil: {valid_leads}/{total_leads}",
    )

    _LEAD_HEADERS = []
    if show_decl_hand:
        _LEAD_HEADERS.append('Spilfører')
    _LEAD_HEADERS.append('Lead type')
    if show_decl_hand:
        _LEAD_HEADERS.append('Udspiller')
    _LEAD_HEADERS.append('Udspil')
    if show_contract_pool:
        _LEAD_HEADERS.append('Farve')
    _LEAD_HEADERS.extend(['Antal', 'Avg Pct Defense', 'Avg Pct Decl', 'Make-rate'])
    lead_header_pos = {h: idx for idx, h in enumerate(_LEAD_HEADERS)}
    lead_color_col_rel = lead_header_pos.get('Farve')
    lead_udspiller_col_rel = lead_header_pos.get('Udspiller')
    lead_udspil_col_rel = lead_header_pos.get('Udspil')

    def _normalize_lead_token(value) -> str | None:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None
        s = str(value).replace('\xa0', '').replace(' ', '').strip().upper()
        return s if s else None

    def _normalize_contract_color_token(value) -> str | None:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None
        s = str(value).strip().upper().replace('UT', 'NT')
        if not s or s == 'UKENDT':
            return None
        if s == 'NT':
            return 'NT'
        if s in ('♠', 'S'):
            return '♠'
        if s in ('♥', 'H'):
            return '♥'
        if s in ('♦', 'D'):
            return '♦'
        if s in ('♣', 'C'):
            return '♣'
        return None

    target_pair_lead_norm = _normalize_lead_token(target_pair_lead_value)
    target_pair_contract_color_norm = _normalize_contract_color_token(
        target_pair_contract_color
    )

    _LEAD_HDR_ROW = _LEAD_START_ROW + 2
    for idx, header in enumerate(_LEAD_HEADERS):
        c = ws.cell(row=_LEAD_HDR_ROW, column=_LEAD_START_COL + idx, value=header)
        _apply_header_style(c)

    for ridx, item in enumerate(lead_summary_rows):
        row_no = _LEAD_HDR_ROW + 1 + ridx
        values = []
        aligns = []
        row_udspiller_should_bold = False

        if show_decl_hand:
            values.append(item.get('decl_hand'))
            aligns.append('center')

        values.append(item.get('lead_type'))
        aligns.append('left')

        if show_decl_hand:
            lead_hand = item.get('lead_hand')
            values.append(lead_hand)
            aligns.append('center')
            lead_hand_norm = _normalize_compass(lead_hand)
            row_udspiller_should_bold = (
                lead_hand_norm is not None and lead_hand_norm in target_pair_dirs
            )

        values.append(item.get('lead_values'))
        aligns.append('left')

        if show_contract_pool:
            values.append(item.get('contract_color'))
            aligns.append('left')

        values.extend([
            item.get('n'),
            None if item.get('avg_pct_defense') is None or pd.isna(item.get('avg_pct_defense')) else round(float(item['avg_pct_defense']), 1),
            None if item.get('avg_pct_decl') is None or pd.isna(item.get('avg_pct_decl')) else round(float(item['avg_pct_decl']), 1),
            None if item.get('make_rate') is None or pd.isna(item.get('make_rate')) else round(float(item['make_rate']) * 100.0, 1),
        ])
        aligns.extend(['right', 'right', 'right', 'right'])

        row_contains_target_lead = False
        if target_pair_lead_norm is not None:
            leads_txt = item.get('lead_values')
            if leads_txt is not None:
                lead_tokens = [
                    _normalize_lead_token(part)
                    for part in str(leads_txt).split(',')
                ]
                row_contains_target_lead = target_pair_lead_norm in {
                    tok for tok in lead_tokens if tok is not None
                }

        row_contract_matches = True
        if show_contract_pool:
            row_contract_color_norm = _normalize_contract_color_token(
                item.get('contract_color')
            )
            row_contract_matches = (
                target_pair_contract_color_norm is not None
                and row_contract_color_norm == target_pair_contract_color_norm
            )

        row_fill = _HILITE_YELLOW if (row_contains_target_lead and row_contract_matches) else None

        for cidx, (val, align) in enumerate(zip(values, aligns)):
            dc = ws.cell(row=row_no, column=_LEAD_START_COL + cidx)
            if lead_udspil_col_rel is not None and cidx == lead_udspil_col_rel:
                _write_with_red_suits(dc, val)
            else:
                dc.value = val

            _apply_data_style(dc, align=align, fill_color=row_fill)
            if (
                lead_udspiller_col_rel is not None
                and cidx == lead_udspiller_col_rel
                and row_udspiller_should_bold
                and Font is not None
            ):
                dc.font = Font(bold=True, color='000000')
            if (
                lead_color_col_rel is not None
                and cidx == lead_color_col_rel
                and val in ('♥', '♦')
                and Font is not None
            ):
                dc.font = Font(color='FF0000')

    pct_rows = [
        r for r in lead_summary_rows
        if r.get('avg_pct_defense') is not None and not pd.isna(r.get('avg_pct_defense'))
    ]
    if len(pct_rows) >= 2:
        def _summary_label(item: dict) -> str:
            parts = []
            if show_decl_hand:
                decl_hand = item.get('decl_hand')
                if decl_hand is not None:
                    parts.append(str(decl_hand))
            parts.append(str(item.get('lead_type', 'ukendt')))
            if show_contract_pool:
                pool = item.get('contract_color')
                if pool is not None:
                    parts.append(str(pool))
            return ' / '.join(parts)

        best = max(pct_rows, key=lambda x: float(x['avg_pct_defense']))
        worst = min(pct_rows, key=lambda x: float(x['avg_pct_defense']))
        msg = (
            f"Indtryk: bedst {_summary_label(best)} ({float(best['avg_pct_defense']):.1f}% defense) | "
            f"svagest {_summary_label(worst)} ({float(worst['avg_pct_defense']):.1f}% defense)"
        )
        info_row = _LEAD_HDR_ROW + 1 + len(lead_summary_rows) + 1
        info_cell = ws.cell(row=info_row, column=_LEAD_START_COL, value=msg)
        if Font is not None:
            info_cell.font = Font(italic=True)


def write_last_tournament_board_layout_sheets(
    writer,
    df: pd.DataFrame,
    per_name: str,
    board_start: int = 1,
    board_end: int = 24,
) -> None:
    """
    Write one layout sheet per board for the latest tournament.

    Default creates sheets for board 1..24 with names:
    - Board1_LastTournament
    - Board2_LastTournament
    - ...
    - Board24_LastTournament
    """
    start = int(board_start)
    end = int(board_end)
    if start > end:
        start, end = end, start

    for board_no in range(start, end + 1):
        sheet_name = (
            "Board1_LastTournament"
            if board_no == 1
            else f"Board{board_no}_LastTournament"
        )
        write_board1_layout_sheet(
            writer,
            df,
            per_name,
            board_no=board_no,
            sheet_name=sheet_name,
        )


def _resolve_row_column(df: pd.DataFrame) -> str | None:
    """Return preferred row/section column name, or None if unavailable."""
    if 'row' in df.columns:
        return 'row'
    if 'section' in df.columns:
        return 'section'
    return None


def _normalize_row_code(value) -> str | None:
    """Normalize row code to uppercase string (A/B/C), or None for missing."""
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    s = str(value).strip().upper()
    return s if s else None


def _normalize_compass(value) -> str | None:
    """Normalize compass value to N/S/Ø/V (accepts E/W aliases)."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    s = str(value).strip().upper()
    if s == 'E':
        s = 'Ø'
    elif s == 'W':
        s = 'V'
    return s if s in ('N', 'S', 'Ø', 'V') else None


def _declarer_side_from_decl(value) -> str | None:
    """Return declarer side (NS/ØV) from declarer direction."""
    d = _normalize_compass(value)
    if d in ('N', 'S'):
        return 'NS'
    if d in ('Ø', 'V'):
        return 'ØV'
    return None


def _defense_side_from_decl(value) -> str | None:
    """Return defense side (NS/ØV) from declarer direction."""
    decl_side = _declarer_side_from_decl(value)
    if decl_side == 'NS':
        return 'ØV'
    if decl_side == 'ØV':
        return 'NS'
    return None


def _to_float_or_none(value) -> float | None:
    """Convert value to float when possible, else None."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


_CONTRACT_TOKEN_RE = re.compile(r'([1-7])\s*(NT|[SHDC♠♥♦♣])', re.IGNORECASE)
_STRAIN_TO_KEY = {
    'S': 'S', '♠': 'S',
    'H': 'H', '♥': 'H',
    'D': 'D', '♦': 'D',
    'C': 'C', '♣': 'C',
    'NT': 'NT', 'UT': 'NT',
}
_STRAIN_DISPLAY = {
    'S': '♠',
    'H': '♥',
    'D': '♦',
    'C': '♣',
    'NT': 'NT',
}


def _to_int_or_none(value) -> int | None:
    """Convert value to int when possible, else None."""
    num = _to_float_or_none(value)
    if num is None or pd.isna(num):
        return None
    try:
        return int(num)
    except (TypeError, ValueError):
        return None


def _normalize_strain_key(value) -> str | None:
    """Normalize strain text/symbol to canonical key (NT/S/H/D/C)."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    s = str(value).strip().upper()
    if not s:
        return None
    return _STRAIN_TO_KEY.get(s)


def _contract_level_and_strain_from_row(row: pd.Series) -> tuple[int | None, str | None]:
    """Extract (level, strain_key) from row fields (level/strain/contract)."""
    level = _to_int_or_none(_row_nonnull(row, 'level'))
    strain_key = _normalize_strain_key(_row_nonnull(row, 'strain'))

    if level is not None and strain_key is not None:
        return level, strain_key

    contract_val = _row_nonnull(row, 'contract', 'contract_raw')
    if contract_val is None:
        return level, strain_key

    txt = str(contract_val).strip().upper().replace('UT', 'NT')
    m = _CONTRACT_TOKEN_RE.search(txt)
    if not m:
        return level, strain_key

    if level is None:
        try:
            level = int(m.group(1))
        except (TypeError, ValueError):
            level = None
    if strain_key is None:
        strain_key = _normalize_strain_key(m.group(2))

    return level, strain_key


def _contract_pool_from_row(row: pd.Series) -> str:
    """Build pooled contract label: 1-5 by strain, while slams stay separate."""
    level, strain_key = _contract_level_and_strain_from_row(row)
    if level is None or strain_key is None:
        return 'ukendt'

    strain_display = _STRAIN_DISPLAY.get(strain_key, strain_key)
    if level >= 6:
        return f"{level}{strain_display}"
    if 1 <= level <= 5:
        return f"1-5{strain_display}"
    return 'ukendt'


def _contract_pool_to_color_symbol(value) -> str:
    """Convert contract-pool text to a strain symbol (or NT)."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 'ukendt'

    s = str(value).strip().upper()
    if not s or s == 'UKENDT':
        return 'ukendt'

    if s.endswith('NT') or s.endswith('UT') or s in ('NT', 'UT'):
        return _STRAIN_DISPLAY['NT']

    # Most contract-pool values end with suit symbol (e.g. 1-5♥, 6♠).
    last = s[-1]
    if last == '♠' or last == 'S':
        return _STRAIN_DISPLAY['S']
    if last == '♥' or last == 'H':
        return _STRAIN_DISPLAY['H']
    if last == '♦' or last == 'D':
        return _STRAIN_DISPLAY['D']
    if last == '♣' or last == 'C':
        return _STRAIN_DISPLAY['C']

    # Fallback for unexpected formats containing suit text.
    if '♠' in s:
        return _STRAIN_DISPLAY['S']
    if '♥' in s:
        return _STRAIN_DISPLAY['H']
    if '♦' in s:
        return _STRAIN_DISPLAY['D']
    if '♣' in s:
        return _STRAIN_DISPLAY['C']

    return 'ukendt'


def _collect_concrete_leads(values: pd.Series) -> str | None:
    """Return unique concrete lead values as comma-separated text."""
    seen: list[str] = []
    for val in values:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            continue
        txt = str(val).strip()
        if not txt or txt.lower() == 'nan':
            continue
        if txt not in seen:
            seen.append(txt)
    if not seen:
        return None
    return ', '.join(seen)


def _decl_hand_from_row(row: pd.Series) -> str:
    """Return declarer hand (N/S/Ø/V) or 'ukendt'."""
    d = _normalize_compass(_row_nonnull(row, 'decl'))
    return d if d is not None else 'ukendt'


def _lead_hand_from_decl_hand(value) -> str:
    """Return opening-lead hand (left of declarer) as N/S/Ø/V or 'ukendt'."""
    d = _normalize_compass(value)
    if d is None:
        return 'ukendt'
    left_of_decl = {
        'N': 'Ø',
        'Ø': 'S',
        'S': 'V',
        'V': 'N',
    }
    return left_of_decl.get(d, 'ukendt')


def _pct_for_side(row: pd.Series, side: str | None) -> float | None:
    """Return pct value for given side from a row."""
    if side == 'NS':
        return _to_float_or_none(_row_nonnull(row, 'pct_NS', 'pct_ns'))
    if side == 'ØV':
        return _to_float_or_none(_row_nonnull(row, 'pct_ØV', 'pct_EW', 'pct_ew'))
    return None


def _pct_defense_from_row(row: pd.Series) -> float | None:
    """Return pct from defense perspective (leader + partner side)."""
    return _pct_for_side(row, _defense_side_from_decl(_row_nonnull(row, 'decl')))


def _pct_decl_from_row(row: pd.Series) -> float | None:
    """Return pct from declarer/dummy perspective."""
    return _pct_for_side(row, _declarer_side_from_decl(_row_nonnull(row, 'decl')))


def _lead_bool(value) -> bool | None:
    """Convert mixed bool-like values to bool/None."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, str):
        s = value.strip().lower()
        if s in ('true', '1', 'yes', 'ja'):
            return True
        if s in ('false', '0', 'no', 'nej'):
            return False
    try:
        return bool(value)
    except Exception:
        return None


def _row_nonnull(row: pd.Series, *candidates):
    """Return first non-null candidate value from a pandas row."""
    for c in candidates:
        if c not in row.index:
            continue
        val = row.get(c)
        if val is not None and not (isinstance(val, float) and pd.isna(val)):
            return val
    return None


def _lead_type_text_from_row(row: pd.Series) -> str:
    """Build textual lead-type description from lead-analysis fields."""
    strategic = _row_nonnull(row, 'lead_strategic_class')
    rank_class = _row_nonnull(row, 'lead_rank_class')
    lead_valid = _lead_bool(_row_nonnull(row, 'lead_valid'))
    excluded = _lead_bool(_row_nonnull(row, 'exclude_from_lead_stats'))

    strategic_txt = str(strategic).strip() if strategic is not None else ''
    rank_txt = str(rank_class).strip() if rank_class is not None else ''

    if strategic_txt:
        if rank_txt and rank_txt.lower() != 'unclear':
            return f"{strategic_txt} ({rank_txt})"
        return strategic_txt

    if rank_txt:
        return rank_txt

    if lead_valid is False or excluded is True:
        return 'ugyldigt lead'

    return 'ukendt'


def _filter_valid_leads(df: pd.DataFrame) -> pd.DataFrame:
    """Filter to rows considered valid for lead statistics."""
    if df.empty:
        return df.copy()

    if 'exclude_from_lead_stats' in df.columns:
        valid_mask = ~df['exclude_from_lead_stats'].apply(lambda v: _lead_bool(v) is True)
    elif 'lead_valid' in df.columns:
        valid_mask = df['lead_valid'].apply(lambda v: _lead_bool(v) is True)
    elif 'lead' in df.columns:
        valid_mask = df['lead'].apply(lambda v: v is not None and str(v).strip() != '')
    else:
        valid_mask = pd.Series(False, index=df.index)

    return df[valid_mask].copy()


def _make_lead_effect_summary_rows(
    df: pd.DataFrame,
    *,
    contract_top_n: int | None = None,
    include_decl_hand: bool = False,
) -> list[dict]:
    """Aggregate lead-effect metrics by lead type and optional contract context."""
    if df.empty:
        return []

    work = df.copy()
    work['_lead_type'] = work.apply(_lead_type_text_from_row, axis=1)
    work['_pct_defense'] = work.apply(_pct_defense_from_row, axis=1)
    work['_pct_decl'] = work.apply(_pct_decl_from_row, axis=1)

    if 'contract_required_tricks' in work.columns:
        req = pd.to_numeric(work['contract_required_tricks'], errors='coerce')
    else:
        if 'level' in work.columns:
            level_num = pd.to_numeric(work['level'], errors='coerce')
        else:
            level_num = pd.Series(index=work.index, dtype=float)
        req = level_num + 6

    if 'tricks' in work.columns:
        tricks_num = pd.to_numeric(work['tricks'], errors='coerce')
    else:
        tricks_num = pd.Series(index=work.index, dtype=float)

    made = (tricks_num >= req)
    made = made.where(~(tricks_num.isna() | req.isna()), other=pd.NA)
    work['_made_num'] = pd.to_numeric(made, errors='coerce')

    use_contract_pool = contract_top_n is not None and int(contract_top_n) > 0
    contract_count_map: dict[str, int] = {}
    contract_rank_map: dict[str, int] = {}

    group_cols: list[str] = []
    if use_contract_pool:
        work['_contract_pool'] = work.apply(_contract_pool_from_row, axis=1)

        pool_counts = (
            work.groupby('_contract_pool', dropna=False)
            .size()
            .reset_index(name='contract_pool_n')
            .sort_values(by=['contract_pool_n', '_contract_pool'], ascending=[False, True])
            .reset_index(drop=True)
        )
        top_n = int(contract_top_n)
        top_pools = pool_counts.head(top_n)['_contract_pool'].astype(str).tolist()

        if not top_pools:
            return []

        contract_count_map = {
            str(r['_contract_pool']): int(r['contract_pool_n'])
            for _, r in pool_counts.iterrows()
        }
        contract_rank_map = {pool: idx + 1 for idx, pool in enumerate(top_pools)}

        work['_contract_pool'] = work['_contract_pool'].astype(str)
        work = work[work['_contract_pool'].isin(top_pools)].copy()
        work['_contract_pool_n'] = work['_contract_pool'].map(contract_count_map)
        work['_contract_pool_rank'] = work['_contract_pool'].map(contract_rank_map)
        group_cols.append('_contract_pool')

    if include_decl_hand:
        work['_decl_hand'] = work.apply(_decl_hand_from_row, axis=1)
        group_cols.append('_decl_hand')

    group_cols.append('_lead_type')

    rows = []
    has_contract_pool = '_contract_pool' in group_cols
    has_decl_hand = '_decl_hand' in group_cols

    for group_key, g in work.groupby(group_cols, dropna=False):
        keys = group_key if isinstance(group_key, tuple) else (group_key,)
        key_idx = 0

        contract_pool = None
        if has_contract_pool:
            contract_pool = str(keys[key_idx])
            key_idx += 1

        decl_hand = None
        if has_decl_hand:
            decl_hand = str(keys[key_idx])
            key_idx += 1

        lead_type = keys[key_idx]

        pct_def_series = pd.to_numeric(g['_pct_defense'], errors='coerce')
        pct_decl_series = pd.to_numeric(g['_pct_decl'], errors='coerce')
        precision_series = (
            pd.to_numeric(g['play_precision_dd'], errors='coerce')
            if 'play_precision_dd' in g.columns
            else pd.Series(dtype=float)
        )
        made_series = pd.to_numeric(g['_made_num'], errors='coerce')
        if 'board_no' in g.columns:
            n_boards = int(pd.to_numeric(g['board_no'], errors='coerce').nunique())
        else:
            n_boards = 0

        item = {
            'lead_type': str(lead_type) if lead_type is not None else 'ukendt',
            'lead_values': _collect_concrete_leads(g['lead']) if 'lead' in g.columns else None,
            'n': len(g),
            'n_boards': n_boards,
            'avg_pct_defense': pct_def_series.mean() if len(pct_def_series) else None,
            'avg_pct_decl': pct_decl_series.mean() if len(pct_decl_series) else None,
            'avg_play_precision_dd': precision_series.mean() if len(precision_series) else None,
            'make_rate': made_series.mean() if len(made_series) else None,
        }

        if has_contract_pool and contract_pool is not None:
            item['contract_pool'] = contract_pool
            item['contract_color'] = _contract_pool_to_color_symbol(contract_pool)
            item['contract_pool_n'] = int(contract_count_map.get(contract_pool, len(g)))
            item['contract_pool_rank'] = int(contract_rank_map.get(contract_pool, 999999))

        if has_decl_hand and decl_hand is not None:
            item['decl_hand'] = decl_hand
            item['lead_hand'] = _lead_hand_from_decl_hand(decl_hand)

        rows.append(item)

    def _desc_sort_or_inf(value) -> float:
        if value is None or pd.isna(value):
            return float('inf')
        return -float(value)

    decl_order = {'N': 1, 'S': 2, 'Ø': 3, 'V': 4, 'ukendt': 5}

    if has_contract_pool:
        rows.sort(
            key=lambda x: (
                int(x.get('contract_pool_rank', 999999)),
                decl_order.get(str(x.get('decl_hand', 'ukendt')), 99) if has_decl_hand else 0,
                _desc_sort_or_inf(x.get('avg_pct_defense')),
                _desc_sort_or_inf(x.get('make_rate')),
                -(x['n'] if x.get('n') is not None else 0),
                str(x.get('lead_type', '')),
            )
        )
    elif has_decl_hand:
        rows.sort(
            key=lambda x: (
                decl_order.get(str(x.get('decl_hand', 'ukendt')), 99),
                _desc_sort_or_inf(x.get('avg_pct_defense')),
                _desc_sort_or_inf(x.get('make_rate')),
                -(x['n'] if x.get('n') is not None else 0),
                str(x.get('lead_type', '')),
            )
        )
    else:
        rows.sort(
            key=lambda x: (
                _desc_sort_or_inf(x.get('avg_pct_defense')),
                _desc_sort_or_inf(x.get('make_rate')),
                -(x['n'] if x.get('n') is not None else 0),
                str(x.get('lead_type', '')),
            )
        )

    return rows


def make_latest_tournament_lead_effect_allboards(
    df: pd.DataFrame,
    rows: tuple[str, ...] = ('A', 'B', 'C'),
    board_start: int = 1,
    board_end: int = 24,
    contract_top_n: int | None = None,
    include_decl_hand: bool = False,
) -> pd.DataFrame:
    """
    Build pooled lead-effect table across all checked boards in latest tournament.

    Output is sorted best → worst by `avg_pct_defense` (high to low).
    """
    start = int(board_start)
    end = int(board_end)
    if start > end:
        start, end = end, start

    if df is None or df.empty:
        return pd.DataFrame()
    if 'tournament_date' not in df.columns or 'board_no' not in df.columns:
        return pd.DataFrame()

    latest_date = df['tournament_date'].max()
    work = df[df['tournament_date'] == latest_date].copy()
    work['_board_no_int'] = pd.to_numeric(work['board_no'], errors='coerce')
    work = work[work['_board_no_int'].between(start, end, inclusive='both')].copy()

    target_rows: list[str] = []
    for r in rows:
        norm = _normalize_row_code(r)
        if norm and norm not in target_rows:
            target_rows.append(norm)
    if not target_rows:
        target_rows = ['A', 'B', 'C']

    row_col = _resolve_row_column(work)
    rows_present_txt = '+'.join(target_rows)
    if row_col is not None:
        work['_row_code'] = work[row_col].apply(_normalize_row_code)
        work = work[work['_row_code'].isin(target_rows)].copy()
        rows_present = sorted(work['_row_code'].dropna().unique().tolist())
        if rows_present:
            rows_present_txt = '+'.join(rows_present)

    total_leads = len(work)
    valid = _filter_valid_leads(work)
    valid_leads = len(valid)

    summary_rows = _make_lead_effect_summary_rows(
        valid,
        contract_top_n=contract_top_n,
        include_decl_hand=include_decl_hand,
    )
    if not summary_rows:
        return pd.DataFrame()

    out = pd.DataFrame(summary_rows)
    out.insert(0, 'rank_best_to_worst', range(1, len(out) + 1))
    out['avg_pct_defense'] = pd.to_numeric(out['avg_pct_defense'], errors='coerce').round(1)
    out['avg_pct_decl'] = pd.to_numeric(out['avg_pct_decl'], errors='coerce').round(1)
    out['make_rate_pct'] = (pd.to_numeric(out['make_rate'], errors='coerce') * 100.0).round(1)

    out['tournament_date'] = latest_date
    out['rows_pooled'] = rows_present_txt
    out['board_start'] = start
    out['board_end'] = end
    out['valid_leads'] = valid_leads
    out['total_leads'] = total_leads

    ordered = [
        'rank_best_to_worst', 'contract_pool_rank', 'contract_pool', 'contract_color', 'contract_pool_n', 'decl_hand',
        'lead_type', 'lead_hand', 'lead_values', 'n', 'n_boards',
        'avg_pct_defense', 'avg_pct_decl', 'make_rate_pct',
        'tournament_date', 'rows_pooled',
        'board_start', 'board_end', 'valid_leads', 'total_leads',
    ]
    return out[[c for c in ordered if c in out.columns]].reset_index(drop=True)


def get_latest_tournament_other_rows_results(
    df: pd.DataFrame,
    base_row: str = 'A',
    other_rows: tuple[str, ...] = ('B', 'C'),
) -> pd.DataFrame:
    """
    Return latest-tournament results for rows other than *base_row*.

    Intended for identifying the "other rows" (typically B+C) from the same
    evening as row A.
    """
    if df is None or df.empty:
        return pd.DataFrame()
    if 'tournament_date' not in df.columns:
        return pd.DataFrame()

    row_col = _resolve_row_column(df)
    if row_col is None:
        return pd.DataFrame()

    latest_date = df['tournament_date'].max()
    out = df[df['tournament_date'] == latest_date].copy()

    out['row_code'] = out[row_col].apply(_normalize_row_code)
    base = _normalize_row_code(base_row)
    target_rows = []
    for row_code in other_rows:
        norm = _normalize_row_code(row_code)
        if norm is not None and norm != base and norm not in target_rows:
            target_rows.append(norm)

    if not target_rows:
        return pd.DataFrame()

    out = out[out['row_code'].isin(target_rows)].copy()

    preferred_cols = [
        'tournament_date', 'board_no', 'row_code',
        'row', 'section',
        'ns1', 'ns2', 'ew1', 'ew2',
        'contract', 'decl', 'lead', 'tricks',
        'score_NS', 'score_ØV', 'pct_NS', 'pct_ØV',
        'N_hand', 'S_hand', 'Ø_hand', 'V_hand',
    ]
    cols = [c for c in preferred_cols if c in out.columns]
    if not cols:
        return out.reset_index(drop=True)

    out = out[cols]
    sort_cols = [c for c in ['row_code', 'board_no'] if c in out.columns]
    if sort_cols:
        out = out.sort_values(sort_cols, na_position='last')
    return out.reset_index(drop=True)


def make_latest_tournament_board_consistency_check(
    df: pd.DataFrame,
    rows: tuple[str, ...] = ('A', 'B', 'C'),
    board_start: int = 1,
    board_end: int = 24,
) -> tuple[pd.DataFrame, dict]:
    """
    Check whether boards are identical across rows (A/B/C) in latest tournament.

    Boards are compared via hand-record signature:
        N_hand | S_hand | Ø_hand | V_hand

    Returns
    -------
    tuple[pd.DataFrame, dict]
        - Per-board consistency report
        - Summary metrics
    """
    start = int(board_start)
    end = int(board_end)
    if start > end:
        start, end = end, start

    target_rows: list[str] = []
    for r in rows:
        norm = _normalize_row_code(r)
        if norm and norm not in target_rows:
            target_rows.append(norm)
    if not target_rows:
        target_rows = ['A', 'B', 'C']

    summary = {
        'latest_date': None,
        'row_column': None,
        'rows_checked': ', '.join(target_rows),
        'board_start': start,
        'board_end': end,
        'boards_checked': max(end - start + 1, 0),
        'boards_ok': 0,
        'boards_missing_row': 0,
        'boards_mismatch': 0,
        'boards_row_internal_mismatch': 0,
        'boards_with_all_rows_present': 0,
        'is_consistent': False,
        'error': '',
    }

    report = pd.DataFrame({'board_no': list(range(start, end + 1))})

    if df is None or df.empty:
        summary['error'] = 'Ingen data tilgængelig.'
        report['status'] = 'MISSING_ROW'
        return report, summary

    if 'tournament_date' not in df.columns:
        summary['error'] = "Kolonnen 'tournament_date' mangler."
        report['status'] = 'MISSING_ROW'
        return report, summary

    if 'board_no' not in df.columns:
        summary['error'] = "Kolonnen 'board_no' mangler."
        report['status'] = 'MISSING_ROW'
        return report, summary

    row_col = _resolve_row_column(df)
    if row_col is None:
        summary['error'] = "Kolonne 'row' eller 'section' mangler."
        report['status'] = 'MISSING_ROW'
        return report, summary

    summary['row_column'] = row_col

    latest_date = df['tournament_date'].max()
    summary['latest_date'] = latest_date

    latest = df[df['tournament_date'] == latest_date].copy()
    latest['row_code'] = latest[row_col].apply(_normalize_row_code)
    latest['_board_no_int'] = pd.to_numeric(latest['board_no'], errors='coerce')

    latest = latest[
        latest['row_code'].isin(target_rows)
        & latest['_board_no_int'].between(start, end, inclusive='both')
    ].copy()

    hand_cols = ['N_hand', 'S_hand', 'Ø_hand', 'V_hand']
    missing_hand_cols = [c for c in hand_cols if c not in latest.columns]
    if missing_hand_cols:
        summary['error'] = f"Manglende hånd-kolonner: {', '.join(missing_hand_cols)}"
        report['status'] = 'MISSING_HAND_COLS'
        return report, summary

    # Per-row result counts in latest tournament (within checked board range)
    counts = latest['row_code'].value_counts().to_dict()
    for row_code in target_rows:
        summary[f'{row_code}_result_rows'] = int(counts.get(row_code, 0))

    # Canonical hand signature per result
    latest['_hand_signature'] = latest[hand_cols].fillna('').astype(str).agg('|'.join, axis=1)

    def _mode_or_first(series: pd.Series):
        if series.empty:
            return None
        mode_vals = series.mode(dropna=False)
        if not mode_vals.empty:
            return mode_vals.iloc[0]
        return series.iloc[0]

    grouped = latest.groupby(['row_code', '_board_no_int'], dropna=False).agg(
        n_results=('board_no', 'size'),
        n_unique_signatures=('_hand_signature', 'nunique'),
        signature=('_hand_signature', _mode_or_first),
    ).reset_index()

    if not grouped.empty:
        grouped['_board_no_int'] = grouped['_board_no_int'].astype(int)

    for row_code in target_rows:
        row_df = grouped[grouped['row_code'] == row_code].set_index('_board_no_int')
        report[f'{row_code}_n_results'] = report['board_no'].map(row_df['n_results'])
        report[f'{row_code}_n_unique_signatures'] = report['board_no'].map(row_df['n_unique_signatures'])
        report[f'{row_code}_signature'] = report['board_no'].map(row_df['signature'])
        report[f'{row_code}_present'] = (
            report[f'{row_code}_signature'].notna()
            & (report[f'{row_code}_signature'].astype(str) != '')
        )
        summary[f'{row_code}_boards_present'] = int(report[f'{row_code}_present'].sum())

    base_row = target_rows[0]
    compare_cols = []
    for row_code in target_rows[1:]:
        cmp_col = f'{base_row}_vs_{row_code}_same'
        compare_cols.append(cmp_col)
        report[cmp_col] = np.where(
            report[f'{base_row}_present'] & report[f'{row_code}_present'],
            report[f'{base_row}_signature'] == report[f'{row_code}_signature'],
            pd.NA,
        )

    present_cols = [f'{r}_present' for r in target_rows]
    report['all_rows_present'] = report[present_cols].all(axis=1)

    if compare_cols:
        report['all_rows_equal'] = (
            report['all_rows_present']
            & report[compare_cols].fillna(False).all(axis=1)
        )
    else:
        report['all_rows_equal'] = report['all_rows_present']

    internal_mismatch = pd.Series(False, index=report.index)
    for row_code in target_rows:
        uniq_col = f'{row_code}_n_unique_signatures'
        internal_mismatch = internal_mismatch | (
            report[uniq_col].fillna(0).astype(float) > 1
        )

    report['status'] = 'OK'
    report.loc[~report['all_rows_present'], 'status'] = 'MISSING_ROW'
    report.loc[report['all_rows_present'] & internal_mismatch, 'status'] = 'ROW_INTERNAL_MISMATCH'
    report.loc[
        report['all_rows_present'] & ~internal_mismatch & ~report['all_rows_equal'],
        'status'
    ] = 'MISMATCH'

    summary['boards_with_all_rows_present'] = int(report['all_rows_present'].sum())
    summary['boards_ok'] = int((report['status'] == 'OK').sum())
    summary['boards_missing_row'] = int((report['status'] == 'MISSING_ROW').sum())
    summary['boards_mismatch'] = int((report['status'] == 'MISMATCH').sum())
    summary['boards_row_internal_mismatch'] = int(
        (report['status'] == 'ROW_INTERNAL_MISMATCH').sum()
    )
    summary['is_consistent'] = (
        summary['boards_checked'] > 0
        and summary['boards_ok'] == summary['boards_checked']
    )

    ordered_cols = ['board_no', 'status', 'all_rows_present', 'all_rows_equal']
    ordered_cols.extend(compare_cols)
    for row_code in target_rows:
        ordered_cols.extend([
            f'{row_code}_present',
            f'{row_code}_n_results',
            f'{row_code}_n_unique_signatures',
            f'{row_code}_signature',
        ])

    report = report[[c for c in ordered_cols if c in report.columns]]
    return report, summary


def print_latest_tournament_board_consistency_summary(summary: dict) -> None:
    """Print readable summary from make_latest_tournament_board_consistency_check()."""
    print("\n" + "="*70)
    print("BOARD-KONSISTENS A/B/C (SENESTE TURNERING)")
    print("="*70)

    if summary.get('error'):
        print(f"\nFejl: {summary['error']}")
        print("\n" + "="*70)
        return

    print(f"\nSeneste turnering: {summary.get('latest_date')}")
    print(f"Rækker checket: {summary.get('rows_checked')}")
    print(f"Board-interval: {summary.get('board_start')}–{summary.get('board_end')}")
    print(f"Boards checket: {summary.get('boards_checked')}")
    print(f"Boards OK: {summary.get('boards_ok')}")
    print(f"Boards med manglende række: {summary.get('boards_missing_row')}")
    print(f"Boards med mismatch: {summary.get('boards_mismatch')}")
    print(f"Boards med intern række-mismatch: {summary.get('boards_row_internal_mismatch')}")

    rows_checked = [s.strip() for s in str(summary.get('rows_checked', '')).split(',') if s.strip()]
    if rows_checked:
        print("\nDækning pr. række:")
        for row_code in rows_checked:
            result_rows = summary.get(f'{row_code}_result_rows', 0)
            boards_present = summary.get(f'{row_code}_boards_present', 0)
            print(f"  {row_code}: {result_rows} resultatrækker, {boards_present} boards med hand-record")

    verdict = "JA" if summary.get('is_consistent') else "NEJ"
    print(f"\nEr boards ens på tværs af rækkerne? {verdict}")
    print("\n" + "="*70)


def make_board_review_all_hands(df: pd.DataFrame) -> pd.DataFrame:
    """
    Board Review rapport: ÉN RÆKKE PER HÅND
    
    Viser hver enkelt hånd med Phase 2.1 reference-data OG hånd-features.
    Sorteret efter tournament_date → board_no → row → performance-forskel.
    
    Parameters:
    -----------
    df : pd.DataFrame
        DataFrame med Phase 2.1 kolonner og hånd-features
    
    Returns:
    --------
    pd.DataFrame
        Board review med alle hænder, sorteret efter dato, board og row
    """
    
    if df.empty:
        return df
        
    # === Define performance_cols FIRST ===
    performance_cols = [
        'tournament_date',
        'board_no',
        'row',
        'contract',
        'expected_pct',
        'pct_NS',
        'pct_vs_expected',
        'pct_vs_expected_abs',
        'Board_Type',
        'competitive_flag',
    ]


    # Vælg relevante kolonner
    cols_to_keep = [
        'tournament_date',
        'board_no',
        'row',
        'contract',
        'expected_pct',
        'pct_NS',
        'pct_vs_expected',
        'pct_vs_expected_abs',
        'Board_Type',
        'competitive_flag',
    ]

    hand_cols = [
        'N_hand', 'S_hand', 'Ø_hand', 'V_hand',
        'N_HCP', 'S_HCP', 'Ø_HCP', 'V_HCP', 'NS_HCP', 'ØV_HCP',
        'N_shape', 'S_shape', 'Ø_shape', 'V_shape',
        'N_shape_SHDC', 'S_shape_SHDC', 'Ø_shape_SHDC', 'V_shape_SHDC',
        'N_balanced', 'S_balanced', 'Ø_balanced', 'V_balanced',
        'N_dist_pts_shortage', 'S_dist_pts_shortage', 'Ø_dist_pts_shortage', 'V_dist_pts_shortage',
        'N_LTC_adj', 'S_LTC_adj', 'Ø_LTC_adj', 'V_LTC_adj', 'NS_LTC_adj', 'ØV_LTC_adj',
        'N_controls', 'S_controls', 'Ø_controls', 'V_controls', 'NS_controls', 'ØV_controls',
        'N_aces', 'S_aces', 'Ø_aces', 'V_aces', 'NS_aces', 'ØV_aces',
        'N_kings', 'S_kings', 'Ø_kings', 'V_kings', 'NS_kings', 'ØV_kings',
        'Declarer_Side', 'Declarer_HCP', 'Defense_HCP', 'HCP_diff',
        'Declarer_LTC_adj', 'Defense_LTC_adj', 'LTC_diff',
        'Suit_Index', 'NT_Index',
    ]

    reference_cols = [
        'contract_norm',
        'double_state',
        'decl',
        'level',
        'strain',
        'lead',
        'tricks',
        
        # ✅ RÅDATA HÆNDER
        'N_hand',
        'S_hand',
        'Ø_hand',
        'V_hand',
        
        # ✅ HCP (Høje kort point)
        'N_HCP',
        'S_HCP',
        'Ø_HCP',
        'V_HCP',
        'NS_HCP',
        'ØV_HCP',
        
        # ✅ KORTFORDELING (shape)
        'N_shape',
        'S_shape',
        'Ø_shape',
        'V_shape',
        'N_shape_SHDC',
        'S_shape_SHDC',
        'Ø_shape_SHDC',
        'V_shape_SHDC',
        
        # ✅ BALANCED
        'N_balanced',
        'S_balanced',
        'Ø_balanced',
        'V_balanced',
        
        # ✅ DISTRIBUTION POINTS
        'N_dist_pts_shortage',
        'S_dist_pts_shortage',
        'Ø_dist_pts_shortage',
        'V_dist_pts_shortage',
        
        # ✅ LTC (Losing Trick Count)
        'N_LTC_adj',
        'S_LTC_adj',
        'Ø_LTC_adj',
        'V_LTC_adj',
        'NS_LTC_adj',
        'ØV_LTC_adj',
        
        # ✅ CONTROLS (Aces + Kings)
        'N_controls',
        'S_controls',
        'Ø_controls',
        'V_controls',
        'NS_controls',
        'ØV_controls',
        
        # ✅ ACES & KINGS
        'N_aces',
        'S_aces',
        'Ø_aces',
        'V_aces',
        'N_kings',
        'S_kings',
        'Ø_kings',
        'V_kings',
        
        # ✅ CONTRACT-SIDE METRICS
        'Declarer_Side',
        'Declarer_HCP',
        'Defense_HCP',
        'HCP_diff',
        'Declarer_LTC_adj',
        'Defense_LTC_adj',
        'LTC_diff',
        'Suit_Index',
        'NT_Index',
        
        # ✅ PHASE 2.1 REFERENCE-DATA
        'field_mode_contract',
        'field_mode_count',
        'field_mode_freq',
        'top2_contract_1',
        'top2_contract_2',
        'top2_count_1',
        'top2_count_2',
        'reference_scope',
        'N_section_played',

        # ✅ DEALER / VULNERABILITY
        'dealer',
        'vul',

        # ✅ DOUBLE DUMMY
        'dd_valid',
        'dd_N_NT', 'dd_N_S', 'dd_N_H', 'dd_N_D', 'dd_N_C', 'dd_N_HCP',
        'dd_S_NT', 'dd_S_S', 'dd_S_H', 'dd_S_D', 'dd_S_C', 'dd_S_HCP',
        'dd_Ø_NT', 'dd_Ø_S', 'dd_Ø_H', 'dd_Ø_D', 'dd_Ø_C', 'dd_Ø_HCP',
        'dd_V_NT', 'dd_V_S', 'dd_V_H', 'dd_V_D', 'dd_V_C', 'dd_V_HCP',

        # ✅ PAR
        'par_score',
        'par_contract',
        'par_side',
    ]

    # Filtrer til kolonner som eksisterer (computed cols tilføjes efter copy)
    source_cols = hand_cols + reference_cols + [
        c for c in performance_cols if c not in ('pct_vs_expected', 'pct_vs_expected_abs')
    ]
    cols_available = [col for col in source_cols if col in df.columns]

    report = df[cols_available].copy()

    # Beregn forskel fra expected_pct
    report['pct_vs_expected'] = report['pct_NS'] - report['expected_pct']
    report['pct_vs_expected_abs'] = abs(report['pct_vs_expected'])
    
    # Sorter efter absolutt forskel (største først)
    report = report.sort_values('pct_vs_expected_abs', ascending=False, na_position='last')
    
    # Omarranger kolonner så performance-data er først
    performance_cols = [
        'tournament_date',
        'board_no',
        'row',  # ✅ NY: row letter
        'contract',
        'expected_pct',
        'pct_NS',
        'pct_vs_expected',
        'pct_vs_expected_abs',
        'Board_Type',
        'competitive_flag',
    ]
    
    other_cols = [col for col in cols_available if col not in performance_cols]
    
    report = report[performance_cols + other_cols]
    
    print(f"✓ Board Review (All Hands): {len(report)} rækker")
    
    return report


def make_board_review_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Board Review rapport: ÉN RÆKKE PER BOARD (aggregeret)
    
    Hver board vises med:
    - Antal hænder
    - Gennemsnit pct_NS
    - Expected pct
    - Gennemsnit forskel
    - Board klassifikation
    - Row (A/B/C)
    - Gennemsnit HCP per side
    - Gennemsnit LTC per side
    
    Sorteret efter Board_Type (Split først), så efter forskel.
    
    Parameters:
    -----------
    df : pd.DataFrame
        DataFrame med Phase 2.1 kolonner og hånd-features
    
    Returns:
    --------
    pd.DataFrame
        Én række per board med aggregeret data
    """
    
    if df.empty:
        return df
    
    # Grupper per board
    grouped = df.groupby(['tournament_date', 'board_no']).agg({
    'row': 'first',
    'contract': 'first',
    'contract_norm': 'first',
    'decl': 'first',
    'level': 'first',
    'strain': 'first',
    'field_mode_contract': 'first',
    'field_mode_count': 'first',
    'field_mode_freq': 'first',
    'top2_contract_1': 'first',
    'top2_contract_2': 'first',
    'expected_pct': 'first',
    'pct_NS': 'mean',
    'reference_scope': 'first',
    'N_section_played': 'first',
    'Board_Type': 'first',
    'competitive_flag': 'first',
    # ✅ LEGG TIL DISSE:
    'NS_HCP': 'mean',
    'ØV_HCP': 'mean',
    'NS_LTC_adj': 'mean',
    'ØV_LTC_adj': 'mean',
    'NS_controls': 'mean',
    'ØV_controls': 'mean',
}).reset_index()

    
    # Rename for klarhed
    grouped = grouped.rename(columns={
        'pct_NS': 'avg_pct_NS',
        'contract': 'contract_actual',
        'NS_HCP': 'avg_NS_HCP',
        'ØV_HCP': 'avg_ØV_HCP',
        'NS_LTC_adj': 'avg_NS_LTC_adj',
        'ØV_LTC_adj': 'avg_ØV_LTC_adj',
        'NS_controls': 'avg_NS_controls',
        'ØV_controls': 'avg_ØV_controls',
    })
    
    # Antal hænder per board
    hand_counts = df.groupby(['tournament_date', 'board_no']).size().reset_index(name='num_hands')
    grouped = grouped.merge(hand_counts, on=['tournament_date', 'board_no'])
    
    # Beregn forskel
    grouped['avg_pct_vs_expected'] = grouped['avg_pct_NS'] - grouped['expected_pct']
    grouped['avg_pct_vs_expected_abs'] = abs(grouped['avg_pct_vs_expected'])
    
    # Sortering: Split først, så efter forskel
    sort_priority = {
        'Split': 0,
        'Dominant': 1,
        'Wild': 2,
        'LOW_SAMPLE': 3
    }
    grouped['sort_priority'] = grouped['Board_Type'].map(sort_priority)
    
    grouped = grouped.sort_values(
        ['sort_priority', 'avg_pct_vs_expected_abs'],
        ascending=[True, False],
        na_position='last'
    ).drop('sort_priority', axis=1)
    
    # Omarranger kolonner: performance → hånd-data → reference
    performance_cols = [
        'tournament_date',
        'board_no',
        'row',
        'num_hands',
        'contract_actual',
        'decl',
        'level',
        'strain',
        'expected_pct',
        'avg_pct_NS',
        'avg_pct_vs_expected',
        'avg_pct_vs_expected_abs',
        'Board_Type',
        'competitive_flag',
        'avg_NS_HCP',
        'avg_ØV_HCP',
        'avg_NS_LTC_adj',
        'avg_ØV_LTC_adj',
        'avg_NS_controls',
        'avg_ØV_controls',
    ]
    
    other_cols = [col for col in grouped.columns if col not in performance_cols]
    
    grouped = grouped[performance_cols + other_cols]
    
    print(f"✓ Board Review (Summary): {len(grouped)} boards")
    print(f"  Split boards: {(grouped['Board_Type'] == 'Split').sum()}")
    print(f"  Dominant boards: {(grouped['Board_Type'] == 'Dominant').sum()}")
    print(f"  Wild boards: {(grouped['Board_Type'] == 'Wild').sum()}")
    
    return grouped


def board_review_statistics(df_all_hands: pd.DataFrame, df_summary: pd.DataFrame) -> dict:
    """
    Beregn deskriptiv statistik for Board Review.
    
    Parameters:
    -----------
    df_all_hands : pd.DataFrame
        Output fra make_board_review_all_hands()
    
    df_summary : pd.DataFrame
        Output fra make_board_review_summary()
    
    Returns:
    --------
    dict
        Statistik som kan bruges til print eller rapport
    """
    
    stats = {
        'total_boards': len(df_summary),
        'total_hands': len(df_all_hands),
        
        'split_boards': (df_summary['Board_Type'] == 'Split').sum(),
        'dominant_boards': (df_summary['Board_Type'] == 'Dominant').sum(),
        'wild_boards': (df_summary['Board_Type'] == 'Wild').sum(),
        
        'avg_pct_overall': df_all_hands['pct_NS'].mean(),
        'avg_expected_pct': df_all_hands['expected_pct'].mean(),
        'avg_performance_vs_expected': df_all_hands['pct_vs_expected'].mean(),
        
        'best_board': df_summary.loc[df_summary['avg_pct_vs_expected'].idxmax()] if len(df_summary) > 0 else None,
        'worst_board': df_summary.loc[df_summary['avg_pct_vs_expected'].idxmin()] if len(df_summary) > 0 else None,
        
        'split_boards_avg_vs_expected': df_summary[df_summary['Board_Type'] == 'Split']['avg_pct_vs_expected'].mean(),
        'dominant_boards_avg_vs_expected': df_summary[df_summary['Board_Type'] == 'Dominant']['avg_pct_vs_expected'].mean(),
        'wild_boards_avg_vs_expected': df_summary[df_summary['Board_Type'] == 'Wild']['avg_pct_vs_expected'].mean(),
    }
    
    return stats


def print_board_review_stats(stats: dict) -> None:
    """
    Print Board Review statistik i læselig format.
    
    Parameters:
    -----------
    stats : dict
        Output fra board_review_statistics()
    """
    
    print("\n" + "="*70)
    print("BOARD REVIEW STATISTIK")
    print("="*70)
    
    print(f"\nOVERALL:")
    print(f"  Total boards: {stats['total_boards']}")
    print(f"  Total hands: {stats['total_hands']}")
    print(f"  Avg hands per board: {stats['total_hands'] / max(stats['total_boards'], 1):.1f}")
    
    print(f"\nBOARD TYPES:")
    print(f"  Split boards: {stats['split_boards']}")
    print(f"  Dominant boards: {stats['dominant_boards']}")
    print(f"  Wild boards: {stats['wild_boards']}")
    
    print(f"\nPERFORMANCE:")
    print(f"  Overall avg pct: {stats['avg_pct_overall']:.1f}%")
    print(f"  Expected avg pct: {stats['avg_expected_pct']:.1f}%")
    print(f"  Difference: {stats['avg_performance_vs_expected']:+.1f}%")
    
    print(f"\nPERFORMANCE BY BOARD TYPE:")
    if not pd.isna(stats['split_boards_avg_vs_expected']):
        print(f"  Split boards: {stats['split_boards_avg_vs_expected']:+.1f}%")
    if not pd.isna(stats['dominant_boards_avg_vs_expected']):
        print(f"  Dominant boards: {stats['dominant_boards_avg_vs_expected']:+.1f}%")
    if not pd.isna(stats['wild_boards_avg_vs_expected']):
        print(f"  Wild boards: {stats['wild_boards_avg_vs_expected']:+.1f}%")
    
    if stats['best_board'] is not None:
        best = stats['best_board']
        print(f"\nBEST BOARD:")
        print(f"  {best['tournament_date']} - Board {best['board_no']} (Row {best['row']})")
        print(f"  Performance: {best['avg_pct_vs_expected']:+.1f}% vs expected")
    
    if stats['worst_board'] is not None:
        worst = stats['worst_board']
        print(f"\nWORST BOARD:")
        print(f"  {worst['tournament_date']} - Board {worst['board_no']} (Row {worst['row']})")
        print(f"  Performance: {worst['avg_pct_vs_expected']:+.1f}% vs expected")
    
    print("\n" + "="*70)