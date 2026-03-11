import sys
import argparse
import shutil
from datetime import datetime, timedelta
import pandas as pd
import requests

try:
    from openpyxl.styles import Font
except ImportError:
    Font = None

try:
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    _RED_INLINE_FONT = InlineFont(color='FF0000')
    _RICH_TEXT_AVAILABLE = True
except ImportError:
    CellRichText = None
    TextBlock = None
    InlineFont = None
    _RED_INLINE_FONT = None
    _RICH_TEXT_AVAILABLE = False

from bridge.data_cache import DataCache
from bridge.crawler import get_recent_tournaments
from bridge.scraper import scrape_spilresultater

from bridge.features import add_hand_features

from bridge.analysis import (
    add_roles_and_pct,
    make_role_summary,
    make_tournament_summary,
    make_declarer_list,
    make_evening_role_matrix,
    make_quarterly_summary_with_ci,
    make_pair_field_report,
    make_pair_declarer_report,
)

# ✅ IMPORT PHASE 2.1 REFERENCE-LAG
from bridge.phase21_reference import add_phase21_fields

# ✅ IMPORT BOARD REVIEW
from bridge.board_review import (
    make_board_review_all_hands,
    make_board_review_summary,
    board_review_statistics,
    make_latest_tournament_lead_effect_allboards,
    print_board_review_stats,
    get_latest_tournament_other_rows_results,
    make_latest_tournament_board_consistency_check,
    print_latest_tournament_board_consistency_summary,
    write_last_tournament_board_layout_sheets,
)

# ✅ IMPORT DECLARER ANALYSIS
from bridge.declarer_analysis import (
    make_declarer_analysis,
    make_declarer_risk_report,
    print_declarer_analysis_highlights,
)

# ✅ IMPORT MVP METRICS
from bridge.mvp_metrics import add_mvp_metrics

HENRIK = "Henrik Friis"
PER = "Per Føge Jensen"
REPORT_EVENING_SHEET = "Rapport - Aften"
REPORT_QUARTER_SHEET = "Rapport - Kvartal"

# Use unique filename to avoid Windows file-lock issues
OUTPUT_FILE = f"Henrik_Per_ANALYSE_{datetime.now():%Y%m%d_%H%M}.xlsx"
LATEST_OUTPUT_FILE = "Henrik_Per_ANALYSE_latest.xlsx"


def _apply_red_suit_symbols(cell) -> None:
    """Color only heart/diamond symbols red in a cell value when possible."""
    val = cell.value
    if val is None:
        return

    text = str(val)
    if '♥' not in text and '♦' not in text:
        return

    if not _RICH_TEXT_AVAILABLE:
        if Font is not None:
            cell.font = Font(color='FF0000')
        return

    parts = []
    i = 0
    while i < len(text):
        if text[i] in ('♥', '♦'):
            parts.append(TextBlock(_RED_INLINE_FONT, text[i]))
            i += 1
        else:
            j = i + 1
            while j < len(text) and text[j] not in ('♥', '♦'):
                j += 1
            parts.append(text[i:j])
            i = j

    cell.value = CellRichText(parts)


def _format_report_sheet(ws, width: int = 12) -> None:
    """Set uniform column width and show numeric values with 0 decimals."""
    if ws is None:
        return

    for col_idx in range(1, ws.max_column + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = width

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if isinstance(val, bool):
                continue
            if isinstance(val, (int, float)):
                cell.number_format = '0'


def parse_arguments():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(
        description='Bridge Analytics - Scrap og analyser bridge-turneringer med smart caching'
    )
    
    parser.add_argument(
        '--cutoff',
        type=str,
        help='Cutoff dato (fra denne dato til i dag). Format: YYYY-MM-DD. Default: sidste 7 dage'
    )
    
    parser.add_argument(
        '--from',
        dest='from_date',
        type=str,
        help='Start dato for interval. Format: YYYY-MM-DD'
    )
    
    parser.add_argument(
        '--to',
        dest='to_date',
        type=str,
        help='Slut dato for interval. Format: YYYY-MM-DD'
    )
    
    parser.add_argument(
        '--force-refresh',
        action='store_true',
        help='Tvinger refresh af alle turneringer i perioden'
    )
    
    parser.add_argument(
        '--cache-status',
        action='store_true',
        help='Vis cache status og exit'
    )
    
    parser.add_argument(
        '--clear-cache',
        action='store_true',
        help='Slet hele cache (kræver bekræftelse)'
    )
    
    parser.add_argument(
        '--backup',
        action='store_true',
        help='Lav backup af cache før start'
    )
    
    return parser.parse_args()


def _load_all_cached_rows(cache: DataCache) -> list[dict]:
    """Load all rows from cached tournaments, preserving/setting section labels."""
    rows: list[dict] = []

    tournaments = cache.manifest.get("tournaments", {})
    tids: list[int] = []
    for tid_txt in tournaments.keys():
        try:
            tids.append(int(tid_txt))
        except (TypeError, ValueError):
            continue

    for tid in sorted(tids):
        cached_data = cache.get_cached_tournament(tid)
        if not cached_data:
            continue

        sections = cached_data.get("sections", {})
        if not isinstance(sections, dict):
            continue

        for section_name, section_rows in sections.items():
            if not isinstance(section_rows, list):
                continue

            for raw_row in section_rows:
                if not isinstance(raw_row, dict):
                    continue

                row = dict(raw_row)

                # Keep explicit section if already present, else use map key.
                if not row.get("section"):
                    row["section"] = section_name

                # Normalize date field name for downstream grouping.
                if not row.get("tournament_date") and row.get("date"):
                    row["tournament_date"] = row.get("date")

                rows.append(row)

    return rows


def main():
    args = parse_arguments()
    
    # ==================== SETUP CACHE ====================
    print("🚀 Initialiserer cache system...")
    cache = DataCache(data_dir="data")
    
    # ==================== HANDLE SPECIAL FLAGS ====================
    
    # Show cache status
    if args.cache_status:
        cache.print_cache_status()
        return
    
    # Clear cache
    if args.clear_cache:
        response = input("⚠️  Sikker på du vil slette hele cache? (ja/nej): ")
        if response.lower() in ['ja', 'yes', 'y']:
            cache.clear_cache(confirm=True)
        return
    
    # Create backup
    if args.backup:
        cache.create_backup()
    
    # ==================== PARSE DATE RANGE ====================
    print("Starter crawler + scraper + analyse...")
    
    start_date, end_date = cache.parse_date_range(
        cutoff=args.cutoff,
        from_date=args.from_date,
        to_date=args.to_date
    )
    
    force_refresh = args.force_refresh
    if force_refresh:
        print("🔄 FORCE REFRESH MODE - Scraper alt på tværs af regler")
    
    # ==================== CRAWL BRIDGE.DK ====================
    print(f"📡 Søger turneringer på bridge.dk fra {start_date} til {end_date}...")

    online_mode = True
    tournaments_in_range = []

    try:
        # Crawler finder turneringer fra bridge.dk
        all_tournaments_on_site = get_recent_tournaments(start_date)
    except requests.exceptions.RequestException as exc:
        online_mode = False
        print(f"⚠ Netværksfejl mod bridge.dk: {exc}")
        print("⚠ Fallback: bruger kun lokale cache-data i valgt periode.")
        tournaments_in_range = cache.get_cached_tournaments_in_range(start_date, end_date)
        if not tournaments_in_range:
            print("Ingen cachede turneringer fundet i perioden, og bridge.dk kunne ikke nås.")
            return
        print(f"Antal cachede turneringer i periode [{start_date} - {end_date}]: {len(tournaments_in_range)}")
    else:
        if not all_tournaments_on_site:
            print("Ingen turneringer fundet indenfor perioden.")
            return

        print(f"Antal turneringer fundet på bridge.dk: {len(all_tournaments_on_site)}")

        # Filter turneringer til vores periode
        tournaments_in_range = [
            t for t in all_tournaments_on_site
            if start_date <= t['date'].date() <= end_date
        ]

        print(f"Antal turneringer i periode [{start_date} - {end_date}]: {len(tournaments_in_range)}")
    
    # ==================== BESLUT OM SCRAPING ====================
    print("\n📋 Beslutter hvad skal skrabes...")
    print("="*70)
    
    tournaments_to_scrape = []
    tournaments_to_use_cache = []

    if not online_mode:
        tournaments_to_use_cache = list(tournaments_in_range)
        print("  Offline mode: scraping deaktiveret, indlæser kun cache.")
    else:
        for tournament in tournaments_in_range:
            tournament_id = tournament['tournament_id']
            tournament_date = tournament['date']
            
            # Tjek om bruger specifikt bad om denne periode (hvis --from/--to er brugt)
            user_requested_older = bool(args.from_date and args.to_date)
            
            should_scrape = cache.should_scrape_tournament(
                tournament_id=tournament_id,
                tournament_date=tournament_date,
                force_refresh=force_refresh,
                user_requested_older=user_requested_older
            )
            
            if should_scrape:
                tournaments_to_scrape.append(tournament)
            else:
                tournaments_to_use_cache.append(tournament)
    
    print("="*70)
    print(f"\n📊 Beslutninger:")
    print(f"  🔄 SCRAPE: {len(tournaments_to_scrape)} turneringer")
    print(f"  💾 CACHE: {len(tournaments_to_use_cache)} turneringer")
    
    # ==================== SCRAPE & CACHE ====================
    print("\n🌐 Starter scraping...")
    all_rows = []
    tournaments_scraped = 0
    
    for t_idx, tournament in enumerate(tournaments_to_scrape, 1):
        tournament_id = tournament['tournament_id']
        tdate = tournament['date']
        sections = tournament['sections']
        
        print(f"\n({t_idx}/{len(tournaments_to_scrape)}) SCRAPE: {tdate.date()} – Turnering {tournament_id}")
        print(f"  Sections: {', '.join([s['name'] for s in sections])}")
        
        tournament_rows = []
        tournament_data = {"tournament_id": tournament_id, "date": str(tdate.date()), "sections": {}}
        
        # Scrape hver section
        for section in sections:
            section_name = section['name']
            section_url = section['spilresultater_url']
            
            print(f"  → Scraper section {section_name}: {section_url}")
            
            try:
                rows = scrape_spilresultater(
                    section_url,
                    tdate,
                    include_hands=True,
                    debug_hands=False,
                )
                
                # Tilføj section kolonne
                for row in rows:
                    row['section'] = section_name
                
                print(f"      ✓ {len(rows)} rækker")
                tournament_rows.extend(rows)
                tournament_data["sections"][section_name] = rows
                
            except Exception as e:
                print(f"      !!! Fejl ved scraping: {e}")
        
        # Gem i cache
        if tournament_rows:
            cache.save_tournament_data(
                tournament_id=tournament_id,
                tournament_date=tdate,
                sections=sections,
                data=tournament_data
            )
            all_rows.extend(tournament_rows)
            tournaments_scraped += 1
    
        # ==================== LOAD CACHED DATA ====================
    print(f"\n💾 Indlæser {len(tournaments_to_use_cache)} turneringer fra cache...")
    
    for tournament in tournaments_to_use_cache:
        tournament_id = tournament['tournament_id']
        tdate = tournament['date']
        
        cached_data = cache.get_cached_tournament(tournament_id)
        
        if cached_data:
            print(f"  ✓ Turnering {tournament_id} ({tdate.date()}) - fra cache")
            
            # ✅ KONVERTER strings tilbage til datetime
            # Fordi JSON lagrer alt som strings
            for section_name, rows in cached_data.get("sections", {}).items():
                for row in rows:
                    row['section'] = section_name
                    
                    # Konverter date strings tilbage til date objekter
                    if 'date' in row and isinstance(row['date'], str):
                        try:
                            row['date'] = datetime.strptime(row['date'], '%Y-%m-%d').date()
                        except:
                            pass
                    
                    all_rows.append(row)
        else:
            print(f"  ❌ Turnering {tournament_id} ({tdate.date()}) - FEJL, cache findes ikke")# ==================== PROCESS DATA ====================
    
    if not all_rows:
        print("Ingen data fundet.")
        return
    
    df_all = pd.DataFrame(all_rows)
    
    print(f"\n✓ I alt {len(df_all)} rækker fra {len(tournaments_in_range)} turneringer")
    print(f"  Sections: {df_all['section'].unique().tolist()}")
    print(f"  Scraped: {tournaments_scraped}, fra Cache: {len(tournaments_to_use_cache)}")

    # ✅ IDENTIFICER B/C RESULTATER + CHECK BOARD-KONSISTENS A/B/C
    print("\nChecker board-konsistens på tværs af rækker A/B/C (seneste turnering)...")
    df_other_rows_latest = get_latest_tournament_other_rows_results(
        df_all,
        base_row='A',
        other_rows=('B', 'C'),
    )
    df_board_abc_check, board_abc_summary = make_latest_tournament_board_consistency_check(
        df_all,
        rows=('A', 'B', 'C'),
        board_start=1,
        board_end=24,
    )
    print_latest_tournament_board_consistency_summary(board_abc_summary)
    print(f"  ✓ Resultater i andre rækker (B+C): {len(df_other_rows_latest)}")

    df_board_abc_summary = pd.DataFrame([board_abc_summary])
    
    print("\nTilføjer hånd-features...")
    df_all = add_hand_features(df_all)
    
    # ✅ TILFØJ PHASE 2.1 REFERENCE-LAG
    print("Tilføjer Phase 2.1 reference-lag (fra alle sections A+B+C+D...)...")
    df_all = add_phase21_fields(df_all, n_min=12)
    print("  ✓ Phase 2.1 felt-data beregnet")
    print(f"    - Board Types fundet: {df_all['Board_Type'].value_counts().to_dict()}")
    print(f"    - Split boards (competitive): {df_all['competitive_flag'].sum()}")

    # ✅ TILFØJ MVP METRICS
    print("Tilføjer MVP analyse-metrikker (melding, spilføring, udspil)...")
    df_all = add_mvp_metrics(df_all)
    print("  ✓ MVP metrikker beregnet")

    # ✅ LEAD-EFFEKT PÅ TVÆRS AF ALLE BOARDS (A+B+C)
    print("Genererer pooled lead-effekt (A+B+C, board 1-24)...")
    df_lead_effect_allboards = make_latest_tournament_lead_effect_allboards(
        df_all,
        rows=('A', 'B', 'C'),
        board_start=1,
        board_end=24,
        contract_top_n=5,
        include_decl_hand=True,
    )
    print(f"  ✓ Lead effect rows: {len(df_lead_effect_allboards)}")
    
    # ✅ BOARD REVIEW ANALYSE (kun A-rækken)
    print("\nGenererer Board Review rapporter (kun A-rækken)...")
    df_a_only = df_all[df_all['section'] == 'A'].copy()
    
    if len(df_a_only) == 0:
        print("Ingen data fra A-rækken!")
        return
    
    df_board_review_all = make_board_review_all_hands(df_a_only)
    df_board_review_summary = make_board_review_summary(df_a_only)
    
    # Statistik
    review_stats = board_review_statistics(df_board_review_all, df_board_review_summary)
    print(f"  ✓ Board Review: {len(df_board_review_all)} boards med hand-records")
    print_board_review_stats(review_stats)
    
    # ✅ DECLARER ANALYSIS (kun A-rækken)
    print("\nGenererer Declarer Analysis...")
    df_pair = df_a_only[
        df_a_only.apply(
            lambda r: (HENRIK in [r["ns1"], r["ns2"], r["ew1"], r["ew2"]]) and
                      (PER in [r["ns1"], r["ns2"], r["ew1"], r["ew2"]]),
            axis=1
        )
    ].copy()
    
    if len(df_pair) == 0:
        print("  (Ingen boards hvor Henrik og Per spiller sammen i A-rækken)")
        df_declarer_analysis = pd.DataFrame()
    else:
        # Tilføj roller + pct
        df_pair = add_roles_and_pct(df_pair, henrik=HENRIK, per=PER)
        
        # Lav Declarer Analysis
        df_declarer_analysis = make_declarer_analysis(df_pair, henrik=HENRIK, per=PER)
        
        print(f"  ✓ Declarer Analysis: {len(df_declarer_analysis)} boards")
        
        # Print highlights
        print_declarer_analysis_highlights(df_declarer_analysis, top_n=5)
    
    # ✅ KLASSISKE RAPPORTER
    print("\nGenererer klassiske rapporter...")

    # Build classic reports from all cached tournaments (A-row), not just current date range.
    cached_rows_all = _load_all_cached_rows(cache)
    if cached_rows_all:
        df_classic_source = pd.DataFrame(cached_rows_all)
        if "section" in df_classic_source.columns:
            df_classic_source = df_classic_source[df_classic_source["section"] == "A"].copy()
        if "tournament_date" not in df_classic_source.columns and "date" in df_classic_source.columns:
            df_classic_source["tournament_date"] = df_classic_source["date"]

        unique_dates = (
            df_classic_source["tournament_date"].nunique()
            if "tournament_date" in df_classic_source.columns
            else 0
        )
        print(
            f"  ℹ Klassiske rapporter bruger cache-historik: "
            f"{len(df_classic_source)} A-række rækker på {unique_dates} spilledatoer"
        )
    else:
        df_classic_source = df_a_only.copy()
        print("  ⚠ Ingen cache-historik fundet; bruger kun valgte periode.")

    # Filter til kun boards hvor de spiller sammen
    df_pair_all = df_classic_source[
        df_classic_source.apply(
            lambda r: (HENRIK in [r["ns1"], r["ns2"], r["ew1"], r["ew2"]]) and
                      (PER in [r["ns1"], r["ns2"], r["ew1"], r["ew2"]]),
            axis=1
        )
    ].copy()
    
    if len(df_classic_source) > 0:
        df_classic_all_roles = add_roles_and_pct(df_classic_source, henrik=HENRIK, per=PER)
        df_evening_matrix = make_evening_role_matrix(df_classic_all_roles)
        df_quarterly = make_quarterly_summary_with_ci(df_classic_all_roles)
    else:
        df_classic_all_roles = pd.DataFrame()
        df_evening_matrix = pd.DataFrame()
        df_quarterly = pd.DataFrame()

    if len(df_pair_all) > 0:
        df_pair_all = add_roles_and_pct(df_pair_all, henrik=HENRIK, per=PER)

        df_declarer = make_declarer_list(df_pair_all)
        df_summary = make_role_summary(df_pair_all)
        df_tournament = make_tournament_summary(df_pair_all)

        print(f"  ✓ Klassiske rapporter genereret")
    else:
        df_declarer = pd.DataFrame()
        df_summary = pd.DataFrame()
        df_tournament = pd.DataFrame()
        if df_evening_matrix.empty and df_quarterly.empty:
            print("  (Ingen data til klassiske rapporter)")
        else:
            print("  ✓ Evening/Quarterly rapporter genereret fra historik")
    
    # ✅ FIELD REPORTS
    print("\nGenererer Field Reports...")
    df_field_defense = make_pair_field_report(df_all, min_boards=50)
    df_field_declarer = make_pair_declarer_report(df_all, min_boards=50)
    print(f"  ✓ Field Reports: {len(df_field_defense)} par i defense, {len(df_field_declarer)} par i declarer")
    
    # ✅ SKRIV EXCEL
    print(f"\nSkriver Excel: {OUTPUT_FILE}")
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        # A/B/C board consistency + other-row results (latest tournament)
        if not df_board_abc_check.empty:
            df_board_abc_check.to_excel(writer, sheet_name='Board_ABC_Check', index=False)
        if not df_board_abc_summary.empty:
            df_board_abc_summary.to_excel(writer, sheet_name='Board_ABC_Summary', index=False)
        if not df_other_rows_latest.empty:
            df_other_rows_latest.to_excel(writer, sheet_name='Rows_BC_Results', index=False)

        # Lead effect pooled across all boards (best -> worst)
        if not df_lead_effect_allboards.empty:
            df_lead_export = df_lead_effect_allboards.copy()
            lead_prefix = ['decl_hand', 'lead_type', 'lead_hand', 'lead_values', 'contract_color']
            lead_prefix = [c for c in lead_prefix if c in df_lead_export.columns]
            lead_rest = [
                c for c in df_lead_export.columns
                if c not in lead_prefix and c not in {'contract_pool', 'contract_pool_n', 'avg_play_precision_dd'}
            ]
            df_lead_export = df_lead_export[lead_prefix + lead_rest]

            df_lead_export = df_lead_export.rename(columns={
                'decl_hand': 'Spilfører',
                'lead_type': 'Lead type',
                'lead_hand': 'Udspiller',
                'lead_values': 'Udspil',
                'contract_color': 'Farve',
            })

            df_lead_export.to_excel(writer, sheet_name='Lead_Effect_AllBoards', index=False)

            ws_lead = writer.sheets.get('Lead_Effect_AllBoards')
            if ws_lead is not None:
                farve_col = None
                udspil_col = None
                for col_idx in range(1, ws_lead.max_column + 1):
                    if ws_lead.cell(row=1, column=col_idx).value == 'Farve':
                        farve_col = col_idx
                    if ws_lead.cell(row=1, column=col_idx).value == 'Udspil':
                        udspil_col = col_idx

                for row_idx in range(2, ws_lead.max_row + 1):
                    if farve_col is not None:
                        val = ws_lead.cell(row=row_idx, column=farve_col).value
                        if val in ('♥', '♦'):
                            _apply_red_suit_symbols(ws_lead.cell(row=row_idx, column=farve_col))
                    if udspil_col is not None:
                        _apply_red_suit_symbols(ws_lead.cell(row=row_idx, column=udspil_col))
        else:
            pd.DataFrame([
                {'message': 'Ingen gyldige lead-data fundet for seneste turnering (A+B+C, board 1-24).'}
            ]).to_excel(writer, sheet_name='Lead_Effect_AllBoards', index=False)

        # Board Review
        df_board_review_all.to_excel(writer, sheet_name='Board_Review_All', index=False)
        df_board_review_summary.to_excel(writer, sheet_name='Board_Review_Summary', index=False)
        
        # Declarer Analysis
        if not df_declarer_analysis.empty:
            df_declarer_analysis.to_excel(writer, sheet_name='Declarer_Analysis', index=False)
        
        # Klassiske rapporter
        if not df_declarer.empty:
            df_declarer.to_excel(writer, sheet_name='Declarer_List', index=False)
        if not df_summary.empty:
            df_summary.to_excel(writer, sheet_name='Role_Summary', index=False)
        if not df_tournament.empty:
            df_tournament.to_excel(writer, sheet_name='Tournament_Summary', index=False)
        if not df_evening_matrix.empty:
            df_evening_matrix.to_excel(writer, sheet_name=REPORT_EVENING_SHEET, index=False)
            _format_report_sheet(writer.sheets.get(REPORT_EVENING_SHEET), width=12)
        if not df_quarterly.empty:
            df_quarterly.to_excel(writer, sheet_name=REPORT_QUARTER_SHEET, index=False)
            _format_report_sheet(writer.sheets.get(REPORT_QUARTER_SHEET), width=12)
        
        # Field Reports
        if not df_field_defense.empty:
            df_field_defense.to_excel(writer, sheet_name='Field_Defense', index=False)
        if not df_field_declarer.empty:
            df_field_declarer.to_excel(writer, sheet_name='Field_Declarer', index=False)

        # MVP Metrics (all rows, deduplicated columns)
        mvp_cols = [
            "tournament_date", "board", "section",
            "contract", "level", "strain", "decl",
            "Combined_HCP", "expected_level_hcp", "level_gap_hcp",
            "contract_aggression_hcp",
            "LTC_combined", "expected_tricks_ltc", "contract_required_tricks",
            "ltc_trick_gap", "ltc_soundness_flag",
            "slam_attempted", "slam_hcp_ok", "slam_ltc_ok",
            "dd_tricks_declarer", "play_precision_dd", "contract_hardness_dd",
            "pct_vs_expected",
            "lead_suit", "lead_card",
        ]
        available_mvp = [c for c in mvp_cols if c in df_all.columns]
        df_mvp = df_all[available_mvp]
        df_mvp = df_mvp.loc[:, ~df_mvp.columns.duplicated()]
        df_mvp.to_excel(writer, sheet_name='MVP_Metrics', index=False)

        # Board layouts (1-24) from latest tournament
        write_last_tournament_board_layout_sheets(writer, df_all, PER, board_start=1, board_end=24)

    latest_copy_ok = False
    try:
        shutil.copy2(OUTPUT_FILE, LATEST_OUTPUT_FILE)
        latest_copy_ok = True
        print(f"↪ Fast kopi opdateret: {LATEST_OUTPUT_FILE}")
    except PermissionError:
        print(
            f"⚠ Kunne ikke opdatere {LATEST_OUTPUT_FILE} (filen er sandsynligvis åben i Excel). "
            "Luk filen og kør igen."
        )
    except OSError as exc:
        print(f"⚠ Kunne ikke opdatere fast kopi: {exc}")

    latest_status = "opdateret" if latest_copy_ok else "ikke opdateret"
    print(
        f"✅ Analyse færdig! Output: {OUTPUT_FILE} | "
        f"Fast kopi: {LATEST_OUTPUT_FILE} ({latest_status})"
    )
    
    # ==================== SHOW CACHE STATUS ====================
    cache.print_cache_status()


if __name__ == "__main__":
    main()