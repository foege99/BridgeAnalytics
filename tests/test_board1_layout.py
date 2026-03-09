"""
Tests for write_board1_layout_sheet() in bridge/board_review.py.
"""

import io
import pandas as pd
import pytest

from openpyxl import Workbook
from unittest.mock import MagicMock

from bridge.board_review import (
    write_board1_layout_sheet,
    write_last_tournament_board_layout_sheets,
    _hand_suit_lines,
    _ROTATIONS,
    _both_names_in_df,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

PER = "Per Føge Jensen"
HENRIK = "Henrik Friis"


def _make_writer_mock():
    """Return a minimal pd.ExcelWriter-like mock with a real openpyxl workbook."""
    wb = Workbook()
    # openpyxl creates a default sheet; remove it so the sheet list is clean
    wb.remove(wb.active)
    mock = MagicMock()
    mock.book = wb
    return mock, wb


def _make_df(**overrides):
    """Create a minimal single-row DataFrame for board 1."""
    row = {
        'tournament_date': '2026-01-15',
        'board_no': 1,
        'section': 'A',
        'ns1': PER,
        'ns2': HENRIK,
        'ew1': 'Opp East',
        'ew2': 'Opp West',
        'N_hand': 'AKT7.QJ3.984.AK2',
        'S_hand': '652.A75.KQ72.J54',
        'Ø_hand': 'QJ93.T862.AT.876',
        'V_hand': '84.K94.J653.QT93',
    }
    row.update(overrides)
    return pd.DataFrame([row])


# ---------------------------------------------------------------------------
# Unit tests for _hand_suit_lines
# ---------------------------------------------------------------------------

def test_hand_suit_lines_four_lines():
    lines = _hand_suit_lines('AKT7.QJ3.984.AK2')
    assert len(lines) == 4


def test_hand_suit_lines_suit_symbols():
    lines = _hand_suit_lines('AKT7.QJ3.984.AK2')
    assert lines[0] == '♠ AKT7'
    assert lines[1] == '♥ QJ3'
    assert lines[2] == '♦ 984'
    assert lines[3] == '♣ AK2'


def test_hand_suit_lines_none_returns_dashes():
    lines = _hand_suit_lines(None)
    for line in lines:
        assert '-' in line


# ---------------------------------------------------------------------------
# Unit tests for _ROTATIONS
# ---------------------------------------------------------------------------

def test_rotation_per_south_no_change():
    rot = _ROTATIONS['S']
    assert rot['bottom'] == 'S'
    assert rot['top'] == 'N'
    assert rot['left'] == 'V'
    assert rot['right'] == 'Ø'


def test_rotation_per_north_flipped():
    rot = _ROTATIONS['N']
    assert rot['bottom'] == 'N'
    assert rot['top'] == 'S'
    assert rot['left'] == 'Ø'
    assert rot['right'] == 'V'


def test_rotation_per_east():
    rot = _ROTATIONS['Ø']
    assert rot['bottom'] == 'Ø'
    assert rot['top'] == 'V'
    assert rot['left'] == 'S'
    assert rot['right'] == 'N'


def test_rotation_per_west():
    rot = _ROTATIONS['V']
    assert rot['bottom'] == 'V'
    assert rot['top'] == 'Ø'
    assert rot['left'] == 'N'
    assert rot['right'] == 'S'


# ---------------------------------------------------------------------------
# Integration tests for write_board1_layout_sheet
# ---------------------------------------------------------------------------

def test_sheet_created():
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, _make_df(), PER)
    assert 'Board1_LastTournament' in wb.sheetnames


def test_custom_board_number_sheet_created_and_titled():
    """Function can render other boards than board 1 via board_no argument."""
    df = _make_df(board_no=2)
    writer, wb = _make_writer_mock()

    write_board1_layout_sheet(writer, df, PER, board_no=2, sheet_name='Board2_LastTournament')

    assert 'Board2_LastTournament' in wb.sheetnames
    ws = wb['Board2_LastTournament']
    title = str(ws.cell(row=1, column=2).value)
    assert 'Spil 2' in title


def test_wrapper_writes_multiple_board_tabs_1_to_3():
    """Wrapper should create one sheet per board, including message sheet for missing boards."""
    df_b1 = _make_df(board_no=1)
    df_b2 = _make_df(board_no=2)
    df = pd.concat([df_b1, df_b2], ignore_index=True)

    writer, wb = _make_writer_mock()
    write_last_tournament_board_layout_sheets(writer, df, PER, board_start=1, board_end=3)

    assert 'Board1_LastTournament' in wb.sheetnames
    assert 'Board2_LastTournament' in wb.sheetnames
    assert 'Board3_LastTournament' in wb.sheetnames

    ws_missing = wb['Board3_LastTournament']
    msg = str(ws_missing.cell(row=1, column=1).value)
    assert 'Spil 3' in msg


def test_traveller_has_lead_type_last_column():
    """Traveller table must include a final 'Lead type' column."""
    df = _make_df(
        decl='N',
        lead='♠K',
        lead_valid=True,
        exclude_from_lead_stats=False,
        lead_strategic_class='sequence_lead',
        lead_rank_class='top_of_sequence',
        pct_NS=41.0,
        pct_ØV=59.0,
    )
    writer, wb = _make_writer_mock()

    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert ws.cell(row=1, column=8).value == 'Spilfører'    # H1
    assert ws.cell(row=2, column=8).value == 'N'            # H2
    assert ws.cell(row=1, column=19).value == 'Lead type'   # S1
    assert ws.cell(row=1, column=17).value == 'Pct Defense' # Q1
    assert ws.cell(row=1, column=18).value == 'Pct Decl'    # R1
    assert ws.cell(row=2, column=17).value == 59.0          # defense side (ØV)
    assert ws.cell(row=2, column=18).value == 41.0          # declarer side (NS)
    lead_type = str(ws.cell(row=2, column=19).value)        # S2
    assert 'sequence_lead' in lead_type


def test_traveller_mirrors_missing_score_side():
    """If only one score side is present, the opposite side is mirrored with opposite sign."""
    df = _make_df(
        score_NS=420,
        score_ØV=None,
        point_NS=7,
        point_ØV=3,
    )
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    # K2/L2 = Score NS / Score ØV
    assert ws.cell(row=2, column=11).value == 420
    assert ws.cell(row=2, column=12).value == -420


def test_pooled_lead_effect_section_written_under_board():
    """A pooled lead-effect section (A+B+C) should be written below DD table."""
    df_a = _make_df(
        row='A', section='A',
        decl='N',
        contract='4♥', strain='♥',
        lead='♠K', level=4, tricks=10, pct_NS=40.0, pct_ØV=60.0,
        lead_valid=True, exclude_from_lead_stats=False,
        lead_strategic_class='sequence_lead', lead_rank_class='top_of_sequence',
        contract_required_tricks=10,
    )
    df_b = _make_df(
        row='B', section='B',
        ns1='B NS1', ns2='B NS2', ew1='B EW1', ew2='B EW2',
        decl='N',
        contract='4♥', strain='♥',
        lead='♣7', level=4, tricks=9, pct_NS=53.0, pct_ØV=47.0,
        lead_valid=True, exclude_from_lead_stats=False,
        lead_strategic_class='partner_suit_candidate', lead_rank_class='2nd_highest',
        contract_required_tricks=10,
    )
    df_c = _make_df(
        row='C', section='C',
        ns1='C NS1', ns2='C NS2', ew1='C EW1', ew2='C EW2',
        decl='N',
        contract='4♥', strain='♥',
        lead='♥8', level=4, tricks=8, pct_NS=58.0, pct_ØV=42.0,
        lead_valid=True, exclude_from_lead_stats=False,
        lead_strategic_class='trump_lead', lead_rank_class='3rd_5th',
        contract_required_tricks=10,
    )
    df = pd.concat([df_a, df_b, df_c], ignore_index=True)

    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    title_rows = [
        r for r in range(1, 120)
        if str(ws.cell(row=r, column=5).value or '').startswith('Lead-effekt (pooled')
    ]
    assert title_rows, "Expected pooled lead-effect title in column E"

    title_row = title_rows[0]
    assert 'A+B+C' in str(ws.cell(row=title_row, column=5).value)
    assert ws.cell(row=title_row + 2, column=5).value == 'Spilfører'
    assert ws.cell(row=title_row + 2, column=6).value == 'Lead type'
    assert ws.cell(row=title_row + 2, column=7).value == 'Udspiller'
    assert ws.cell(row=title_row + 2, column=8).value == 'Udspil'
    assert ws.cell(row=title_row + 2, column=9).value == 'Farve'
    assert ws.cell(row=title_row + 3, column=5).value == 'N'
    assert ws.cell(row=title_row + 3, column=7).value == 'Ø'
    assert ws.cell(row=title_row + 3, column=8).value == '♠K'
    assert ws.cell(row=title_row + 3, column=9).value == '♥'

    lead_types = [
        str(ws.cell(row=r, column=6).value)
        for r in range(title_row + 3, title_row + 12)
        if ws.cell(row=r, column=6).value is not None
    ]
    assert any('sequence_lead' in s for s in lead_types)


def test_pooled_lead_effect_sorted_best_to_worst_even_if_count_lower():
    """Lead-effect table under board must rank by performance (best->worst), not frequency."""
    # Best type appears once with high pct
    df_best = _make_df(
        row='A', section='A',
        decl='N',
        lead='♠K', level=4, tricks=10, pct_NS=30.0, pct_ØV=70.0,
        lead_valid=True, exclude_from_lead_stats=False,
        lead_strategic_class='best_type', lead_rank_class='unclear',
        contract_required_tricks=10,
    )

    # Worse type appears three times with lower pct
    df_w1 = _make_df(
        row='B', section='B', ns1='B1', ns2='B2', ew1='B3', ew2='B4',
        decl='N',
        lead='♣7', level=4, tricks=9, pct_NS=58.0, pct_ØV=42.0,
        lead_valid=True, exclude_from_lead_stats=False,
        lead_strategic_class='worse_type', lead_rank_class='unclear',
        contract_required_tricks=10,
    )
    df_w2 = _make_df(
        row='C', section='C', ns1='C1', ns2='C2', ew1='C3', ew2='C4',
        decl='N',
        lead='♣7', level=4, tricks=9, pct_NS=56.0, pct_ØV=44.0,
        lead_valid=True, exclude_from_lead_stats=False,
        lead_strategic_class='worse_type', lead_rank_class='unclear',
        contract_required_tricks=10,
    )
    df_w3 = _make_df(
        row='A', section='A', ns1='A3', ns2='A4', ew1='A5', ew2='A6',
        decl='N',
        lead='♣7', level=4, tricks=9, pct_NS=59.0, pct_ØV=41.0,
        lead_valid=True, exclude_from_lead_stats=False,
        lead_strategic_class='worse_type', lead_rank_class='unclear',
        contract_required_tricks=10,
    )

    df = pd.concat([df_best, df_w1, df_w2, df_w3], ignore_index=True)

    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    title_row = next(
        r for r in range(1, 180)
        if str(ws.cell(row=r, column=5).value or '').startswith('Lead-effekt (pooled')
    )

    # First data row is title+3 (header at title+2)
    first_lead_type = str(ws.cell(row=title_row + 3, column=6).value)
    assert 'best_type' in first_lead_type


def test_pooled_lead_effect_highlights_target_pair_lead_row():
    """Lead-effect row containing the actual target-pair lead should be highlighted yellow."""
    df_a = _make_df(
        row='A', section='A',
        ns1=HENRIK, ns2=PER,
        decl='N',
        contract='4♥', strain='♥',
        lead='♥A', level=4, tricks=10, pct_NS=40.0, pct_ØV=60.0,
        lead_valid=True, exclude_from_lead_stats=False,
        lead_strategic_class='sequence_lead', lead_rank_class='top_of_sequence',
        contract_required_tricks=10,
    )
    df_b = _make_df(
        row='B', section='B',
        ns1='B NS1', ns2='B NS2', ew1='B EW1', ew2='B EW2',
        decl='N',
        contract='4♥', strain='♥',
        lead='♣7', level=4, tricks=9, pct_NS=53.0, pct_ØV=47.0,
        lead_valid=True, exclude_from_lead_stats=False,
        lead_strategic_class='partner_suit_candidate', lead_rank_class='2nd_highest',
        contract_required_tricks=10,
    )
    df_c = _make_df(
        row='C', section='C',
        ns1='C NS1', ns2='C NS2', ew1='C EW1', ew2='C EW2',
        decl='N',
        contract='4♥', strain='♥',
        lead='♦8', level=4, tricks=8, pct_NS=58.0, pct_ØV=42.0,
        lead_valid=True, exclude_from_lead_stats=False,
        lead_strategic_class='trump_lead', lead_rank_class='3rd_5th',
        contract_required_tricks=10,
    )
    df = pd.concat([df_a, df_b, df_c], ignore_index=True)

    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    title_row = next(
        r for r in range(1, 180)
        if str(ws.cell(row=r, column=5).value or '').startswith('Lead-effekt (pooled')
    )
    hdr_row = title_row + 2
    udspil_col = 8  # E=5; headers: Spilfører, Lead type, Udspiller, Udspil

    highlighted = False
    for r in range(hdr_row + 1, hdr_row + 20):
        v = ws.cell(row=r, column=udspil_col).value
        if v is None:
            continue
        if '♥A' in str(v):
            rgb = str(ws.cell(row=r, column=udspil_col).fill.fgColor.rgb or '')
            assert rgb.endswith('FFF2CC')
            highlighted = True
    assert highlighted, "Expected at least one highlighted lead row containing ♥A"


def test_pooled_lead_effect_highlight_requires_contract_color_match():
    """Yellow lead-row highlight requires both matching lead and matching contract color."""
    df_target = _make_df(
        row='A', section='A',
        ns1=HENRIK, ns2=PER,
        decl='N',
        contract='3NT', strain='NT', level=3,
        lead='♠K', tricks=9, pct_NS=42.0, pct_ØV=58.0,
        lead_valid=True, exclude_from_lead_stats=False,
        lead_strategic_class='sequence_lead', lead_rank_class='top_of_sequence',
        contract_required_tricks=9,
    )
    # Same lead as target pair, but different contract color (spades).
    df_spade = _make_df(
        row='B', section='B',
        ns1='B NS1', ns2='B NS2', ew1='B EW1', ew2='B EW2',
        decl='N',
        contract='4♠', strain='♠', level=4,
        lead='♠K', tricks=10, pct_NS=55.0, pct_ØV=45.0,
        lead_valid=True, exclude_from_lead_stats=False,
        lead_strategic_class='sequence_lead', lead_rank_class='top_of_sequence',
        contract_required_tricks=10,
    )
    df = pd.concat([df_target, df_spade], ignore_index=True)

    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    title_row = next(
        r for r in range(1, 180)
        if str(ws.cell(row=r, column=5).value or '').startswith('Lead-effekt (pooled')
    )
    hdr_row = title_row + 2

    # E=5; columns: E=Spilfører, F=Lead type, G=Udspiller, H=Udspil, I=Farve
    udspil_col = 8
    farve_col = 9

    row_nt = None
    row_spade = None
    for r in range(hdr_row + 1, hdr_row + 20):
        udspil = ws.cell(row=r, column=udspil_col).value
        farve = ws.cell(row=r, column=farve_col).value
        if udspil is None or farve is None:
            continue
        if '♠K' not in str(udspil):
            continue
        if str(farve) == 'NT':
            row_nt = r
        elif str(farve) == '♠':
            row_spade = r

    assert row_nt is not None
    assert row_spade is not None

    nt_rgb = str(ws.cell(row=row_nt, column=udspil_col).fill.fgColor.rgb or '')
    spade_rgb = str(ws.cell(row=row_spade, column=udspil_col).fill.fgColor.rgb or '')
    assert nt_rgb.endswith('FFF2CC')
    assert not spade_rgb.endswith('FFF2CC')


def test_pooled_lead_effect_udspiller_bold_when_target_pair_leads():
    """Udspiller direction letter should be bold when the opening leader is Henrik/Per."""
    df = _make_df(
        row='A', section='A',
        ns1=HENRIK, ns2=PER,
        # decl=Ø => opening leader is S (left of declarer), i.e. target pair.
        decl='Ø',
        contract='3NT', strain='NT', level=3,
        lead='♠K', tricks=9, pct_NS=45.0, pct_ØV=55.0,
        lead_valid=True, exclude_from_lead_stats=False,
        lead_strategic_class='sequence_lead', lead_rank_class='top_of_sequence',
        contract_required_tricks=9,
    )

    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    title_row = next(
        r for r in range(1, 180)
        if str(ws.cell(row=r, column=5).value or '').startswith('Lead-effekt (pooled')
    )
    hdr_row = title_row + 2

    # E=5; columns: E=Spilfører, F=Lead type, G=Udspiller
    udspiller_col = 7
    first_data_row = hdr_row + 1

    assert ws.cell(row=first_data_row, column=5).value == 'Ø'
    assert ws.cell(row=first_data_row, column=udspiller_col).value == 'S'
    assert ws.cell(row=first_data_row, column=udspiller_col).font.bold is True


def test_per_at_bottom_when_north():
    """Per is ns1 (North) → should appear at bottom with label 'N: Per Føge Jensen'."""
    df = _make_df()  # ns1 = Per = N
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    # Bottom player cell is row 14, col 2
    bottom_cell = ws.cell(row=14, column=2).value
    assert bottom_cell is not None
    assert PER in bottom_cell
    assert bottom_cell.startswith('N:')


def test_per_at_bottom_when_south():
    """Per is ns2 (South) → should appear at bottom with label 'S: Per Føge Jensen'."""
    df = _make_df(ns1=HENRIK, ns2=PER)
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    bottom_cell = ws.cell(row=14, column=2).value
    assert bottom_cell is not None
    assert PER in bottom_cell
    assert bottom_cell.startswith('S:')


def test_per_at_bottom_when_east():
    """Per is ew1 (East/Ø) → should appear at bottom with label 'Ø: Per Føge Jensen'."""
    df = _make_df(ns1='Player A', ns2='Player B', ew1=PER, ew2=HENRIK)
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    bottom_cell = ws.cell(row=14, column=2).value
    assert bottom_cell is not None
    assert PER in bottom_cell
    assert bottom_cell.startswith('Ø:')


def test_per_at_bottom_when_west():
    """Per is ew2 (West/V) → should appear at bottom with label 'V: Per Føge Jensen'."""
    df = _make_df(ns1='Player A', ns2='Player B', ew1=HENRIK, ew2=PER)
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    bottom_cell = ws.cell(row=14, column=2).value
    assert bottom_cell is not None
    assert PER in bottom_cell
    assert bottom_cell.startswith('V:')


def test_hand_written_as_four_suit_lines():
    """The bottom hand must be spread over 4 rows (rows 15-18, col 2)."""
    df = _make_df()  # Per = N → N_hand = 'AKT7.QJ3.984.AK2'
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    lines = [ws.cell(row=15 + i, column=2).value for i in range(4)]
    assert str(lines[0]) == '♠ AKT7'
    # ♥ and ♦ may be rich text (red symbol); str() normalises both cases
    assert str(lines[1]) == '♥ QJ3'
    assert str(lines[2]) == '♦ 984'
    assert str(lines[3]) == '♣ AK2'


def test_latest_date_used():
    """When multiple dates exist, only the latest date is used."""
    df_old = _make_df(tournament_date='2025-12-01', ns1='Old Player', ns2='Old 2',
                      ew1='Old 3', ew2='Old 4')
    df_new = _make_df(tournament_date='2026-01-15')
    df = pd.concat([df_old, df_new], ignore_index=True)

    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    # Title cell (row 1 col 2) should reference the latest date
    title = ws.cell(row=1, column=2).value
    assert '2026-01-15' in str(title)
    assert '2025-12-01' not in str(title)


def test_missing_board1_writes_message():
    """When board 1 is absent from the latest tournament, a message is written."""
    df = _make_df(board_no=5)  # no board 1
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    msg = ws.cell(row=1, column=1).value
    assert msg is not None
    assert 'Spil 1' in msg


def test_per_not_found_writes_message():
    """When Per is not in board 1, a descriptive message is written."""
    df = _make_df(ns1='Someone', ns2='Else', ew1='Another', ew2='Player')
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    msg = ws.cell(row=1, column=1).value
    assert msg is not None
    assert PER in msg


def test_empty_dataframe_writes_message():
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, pd.DataFrame(), PER)
    ws = wb['Board1_LastTournament']

    msg = ws.cell(row=1, column=1).value
    assert msg is not None


def test_missing_board_no_column_writes_message():
    df = pd.DataFrame([{'tournament_date': '2026-01-15', 'ns1': PER}])
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    msg = ws.cell(row=1, column=1).value
    assert msg is not None


# ---------------------------------------------------------------------------
# Tests for new features: HCP, header block, right-side info, red suits
# ---------------------------------------------------------------------------

def test_player_label_includes_hcp_computed():
    """Player label must include HCP computed from hand (no HCP column supplied)."""
    df = _make_df()  # Per = N, N_hand = 'AKT7.QJ3.984.AK2' → 17 HCP
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    # Bottom cell (Per) row 14, col 2
    label = ws.cell(row=14, column=2).value
    assert '17 HCP' in label


def test_player_label_includes_hcp_from_column():
    """Player label uses per-hand HCP column when present."""
    df = _make_df(N_HCP=21)  # override with explicit column value
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    label = ws.cell(row=14, column=2).value
    assert '21 HCP' in label


def test_header_block_written_col_e():
    """Header block: title/date in B1, dealer in A3, zone in A6."""
    df = _make_df(dealer='N', vul='NS')
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert 'Spil 1' in str(ws.cell(row=1, column=2).value)      # B1
    assert '2026-01-15' in str(ws.cell(row=1, column=2).value)  # B1
    assert 'N' in str(ws.cell(row=3, column=1).value)           # A3 Dealer
    assert 'NS' in str(ws.cell(row=6, column=1).value)          # A6 Zone


def test_header_block_fallback_when_no_dealer_zone():
    """Header block shows '(ukendt)' when dealer/zone columns absent."""
    df = _make_df()  # no dealer or vul column
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert '(ukendt)' in str(ws.cell(row=3, column=1).value)   # A3 Dealer
    assert '(ukendt)' in str(ws.cell(row=6, column=1).value)   # A6 Zone


def test_bid_scaffold_headers_widths_and_default_green():
    """Bid scaffold should reserve A20:D20+ with width 18 and green headers when no side is vulnerable."""
    df = _make_df(vul='-')
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert [ws.cell(row=20, column=c).value for c in range(1, 5)] == ['S', 'V', 'N', 'Ø']
    assert ws.column_dimensions['A'].width == 18
    assert ws.column_dimensions['B'].width == 18
    assert ws.column_dimensions['C'].width == 18
    assert ws.column_dimensions['D'].width == 18

    for c in range(1, 5):
        rgb = str(ws.cell(row=20, column=c).fill.fgColor.rgb or '')
        assert rgb.endswith('EAF2E3')


def test_bid_scaffold_header_zone_coloring_ns_only():
    """When NS are vulnerable, S/N headers are pink and V/Ø stay green."""
    df = _make_df(vul='NS')
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    s_rgb = str(ws.cell(row=20, column=1).fill.fgColor.rgb or '')
    v_rgb = str(ws.cell(row=20, column=2).fill.fgColor.rgb or '')
    n_rgb = str(ws.cell(row=20, column=3).fill.fgColor.rgb or '')
    o_rgb = str(ws.cell(row=20, column=4).fill.fgColor.rgb or '')

    assert s_rgb.endswith('F4E4E8')
    assert n_rgb.endswith('F4E4E8')
    assert v_rgb.endswith('EAF2E3')
    assert o_rgb.endswith('EAF2E3')


def test_bid_scaffold_writes_opening_bid_for_dealer_north():
    """Dealer North with 5+ spades should place opening bid in N column (C21)."""
    df = _make_df(
        dealer='N',
        # 13 HCP, 5-4 majors -> 1S under five-card-major profile.
        N_hand='AKQJ9.8765.3.K2',
    )
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert ws.cell(row=21, column=3).value == '1♠'


def test_bid_scaffold_writes_pas_when_opening_threshold_not_met():
    """Very weak dealer hand should produce PAS in dealer column."""
    df = _make_df(
        dealer='S',
        S_hand='T9842.83.742.953',
    )
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert ws.cell(row=21, column=1).value == 'PAS'


def test_bid_scaffold_writes_opening_bid_for_dealer_east_nt_policy():
    """Dealer East uses EW profile and balanced 17 HCP should open 1NT in Ø column (D21)."""
    df = _make_df(
        dealer='Ø',
        Ø_hand='AKQ2.QJ3.A32.J54',
    )
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert ws.cell(row=21, column=4).value == '1NT'


def test_bid_scaffold_opening_log_block_contains_context_and_choice():
    """Opening decision log should be written below bidding table with a final choice line."""
    df = _make_df(
        dealer='N',
        N_hand='AKQJ9.8765.3.K2',
    )
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert ws.cell(row=36, column=1).value == 'Åbningslog'

    log_lines = [
        str(ws.cell(row=r, column=1).value or '')
        for r in range(37, 47)
    ]
    assert any('Kontekst:' in line for line in log_lines)
    assert any('Valg: 1♠' in line for line in log_lines)


def test_bid_scaffold_opening_log_shows_pas_for_weak_hand():
    """Weak hand should produce PAS and the same final choice in the log block."""
    df = _make_df(
        dealer='S',
        S_hand='T9842.83.742.953',
    )
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    log_lines = [
        str(ws.cell(row=r, column=1).value or '')
        for r in range(37, 47)
    ]
    assert any('Valg: PAS' in line for line in log_lines)


def test_bid_scaffold_second_hand_opens_when_first_hand_passes():
    """If 1st hand passes, 2nd hand should be evaluated as opening seat."""
    df = _make_df(
        dealer='N',
        N_hand='T9842.83.742.953',      # 1H should pass
        Ø_hand='AKQ2.QJ3.A32.J54',      # 2H should open 1NT
    )
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert ws.cell(row=21, column=3).value == 'PAS'   # N
    assert ws.cell(row=21, column=4).value == '1NT'   # Ø


def test_bid_scaffold_second_hand_overcall_is_higher_than_first_hand_bid():
    """2nd hand overcall must be above 1st hand bid (e.g., 1S -> 2H)."""
    df = _make_df(
        dealer='N',
        N_hand='AKQJ9.8765.3.K2',       # 1H opens 1S
        Ø_hand='82.AKQJ97.54.842',      # 2H long hearts -> overcall
    )
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert ws.cell(row=21, column=3).value == '1♠'
    assert str(ws.cell(row=21, column=4).value) == '2♥'


def test_bid_scaffold_second_hand_takeout_double_option():
    """2nd hand can choose takeout double over opponent's one-level suit opening."""
    df = _make_df(
        dealer='N',
        N_hand='KJ4.Q4.AK974.83',       # 1H opens 1D
        Ø_hand='AQJ4.KT93.2.AQ84',      # short diamonds + both majors + 12+ HCP
    )
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert str(ws.cell(row=21, column=3).value) == '1♦'
    assert ws.cell(row=21, column=4).value == 'X'


def test_bid_scaffold_second_hand_wraps_to_next_row_after_column_d():
    """If 1H is in Ø-column (D), 2H must wrap to next row at S-column (A)."""
    df = _make_df(
        dealer='Ø',
        Ø_hand='AKQ2.QJ3.A32.J54',      # 1H opens 1NT in D21
        S_hand='T9842.83.742.953',      # 2H no action -> PAS, must be A22
    )
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert ws.cell(row=21, column=4).value == '1NT'
    assert ws.cell(row=21, column=1).value is None
    assert ws.cell(row=22, column=1).value == 'PAS'


def test_bid_scaffold_third_hand_opens_when_first_two_pass():
    """If 1H and 2H both pass, 3H must be treated as opening seat."""
    df = _make_df(
        dealer='N',
        N_hand='T9842.83.742.953',          # 1H PAS
        Ø_hand='T763.942.J86.874',          # 2H PAS
        S_hand='AKQJ9.8765.3.K2',           # 3H opens 1S
    )
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert ws.cell(row=21, column=3).value == 'PAS'   # N
    assert ws.cell(row=21, column=4).value == 'PAS'   # Ø
    assert ws.cell(row=22, column=1).value == '1♠'    # S


def test_bid_scaffold_third_hand_considers_first_and_second_calls():
    """3H should react to 1H opening and 2H overcall (simple raise in this MVP)."""
    df = _make_df(
        dealer='N',
        N_hand='AKQJ9.8765.3.K2',           # 1H = 1S
        Ø_hand='82.AKQJ97.54.842',          # 2H = 2H overcall
        S_hand='T84.QJ3.A742.KJ5',          # 3H supports spades
    )
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert ws.cell(row=21, column=3).value == '1♠'
    assert str(ws.cell(row=21, column=4).value) == '2♥'
    assert ws.cell(row=22, column=1).value == '2♠'


def test_bid_scaffold_log_lines_start_with_current_bid():
    """Log lines should begin with seat + current bid (e.g., N, 1♠: ...)."""
    df = _make_df(
        dealer='N',
        N_hand='AKQJ9.8765.3.K2',
        Ø_hand='82.AKQJ97.54.842',
        S_hand='T84.QJ3.A742.KJ5',
    )
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    log_lines = [
        str(ws.cell(row=r, column=1).value or '')
        for r in range(37, 57)
    ]

    assert any(line.startswith('N, 1♠:') for line in log_lines)
    assert any(line.startswith('Ø, 2♥:') for line in log_lines)
    assert any(line.startswith('S, 2♠:') for line in log_lines)


def test_bid_scaffold_fourth_hand_bid_and_wrap_after_three_calls():
    """Fourth hand should be added to auction sequence with normal wrap rules."""
    df = _make_df(
        dealer='N',
        N_hand='T9842.83.742.953',          # 1H PAS at C21
        Ø_hand='T763.942.J86.874',          # 2H PAS at D21
        S_hand='AKQJ9.8765.3.K2',           # 3H 1S at A22
        V_hand='82.AKQJ97.54.842',          # 4H 2H at B22
    )
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert ws.cell(row=21, column=3).value == 'PAS'
    assert ws.cell(row=21, column=4).value == 'PAS'
    assert ws.cell(row=22, column=1).value == '1♠'
    assert str(ws.cell(row=22, column=2).value) == '2♥'


def test_bid_scaffold_log_includes_fourth_hand_with_bid_prefix():
    """Log should include fourth-hand lines prefixed with seat and current bid."""
    df = _make_df(
        dealer='N',
        N_hand='T9842.83.742.953',
        Ø_hand='T763.942.J86.874',
        S_hand='AKQJ9.8765.3.K2',
        V_hand='82.AKQJ97.54.842',
    )
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    log_lines = [
        str(ws.cell(row=r, column=1).value or '')
        for r in range(37, 80)
    ]
    assert any(line.startswith('Vest, 2♥:') for line in log_lines)


def test_bid_scaffold_second_round_places_all_four_calls():
    """Second round should place calls 5-8 in sequence with normal row wrap."""
    df = _make_df(
        dealer='N',
        N_hand='JT64.AKJT9.95.Q6',
        Ø_hand='A7.3.AKJ732.T942',
        S_hand='KQ852.Q2.Q64.873',
        V_hand='93.87654.T8.AKJ5',
    )
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    # Round 1: C21, D21, A22, B22
    assert ws.cell(row=21, column=3).value is not None
    assert ws.cell(row=21, column=4).value is not None
    assert ws.cell(row=22, column=1).value is not None
    assert ws.cell(row=22, column=2).value is not None

    # Round 2: C22, D22, A23, B23
    assert ws.cell(row=22, column=3).value is not None
    assert ws.cell(row=22, column=4).value is not None
    assert ws.cell(row=23, column=1).value is not None
    assert ws.cell(row=23, column=2).value is not None


def test_bid_scaffold_log_includes_second_round_side_seat_prefixes():
    """Second-round log lines should also start with seat + bid prefix."""
    df = _make_df(
        dealer='N',
        N_hand='JT64.AKJT9.95.Q6',
        Ø_hand='A7.3.AKJ732.T942',
        S_hand='KQ852.Q2.Q64.873',
        V_hand='93.87654.T8.AKJ5',
    )
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    log_lines = [
        str(ws.cell(row=r, column=1).value or '')
        for r in range(37, 120)
    ]

    # Second-round prefixes expected in this scenario:
    assert any(line.startswith('N, 3♠:') for line in log_lines)
    assert any(line.startswith('Ø, 4♣:') for line in log_lines)
    assert any(line.startswith('S, 4♠:') for line in log_lines)
    assert any(line.startswith('Vest, 5♣:') for line in log_lines)


def test_bid_scaffold_fourth_hand_avoids_enemy_suit_natural_bid():
    """After 1♥ - 2♦ - 2♠, fourth hand should not use enemy ♥ as natural bid."""
    df = _make_df(
        dealer='N',
        N_hand='JT64.AKJT9.95.Q6',       # 1H = 1♥
        Ø_hand='A7.3.AKJ732.T942',       # 2H = 2♦
        S_hand='KQ852.Q2.Q64.873',       # 3H = 2♠
        V_hand='93.87654.T8.AKJ5',       # 4H should avoid 3♥ as natural (cuebid suit)
    )
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert str(ws.cell(row=21, column=3).value) == '1♥'
    assert str(ws.cell(row=21, column=4).value) == '2♦'
    assert ws.cell(row=22, column=1).value == '2♠'
    # Fourth call at B22 must not be 3♥; with this heuristic it becomes 3♣.
    assert str(ws.cell(row=22, column=2).value) != '3♥'
    assert str(ws.cell(row=22, column=2).value) == '3♣'


def test_bid_scaffold_opener_shows_four_card_major_after_minor_support():
    """After 1♣-P-2♣-P, opener with 4 hearts should show 2♥ before raising clubs further."""
    df = _make_df(
        dealer='V',
        V_hand='Q6.QJ87.AK8.8532',
        Ø_hand='A75.KT963.Q.K976',
        N_hand='3.A42.9632.AQJT4',
        S_hand='KJT9842.5.JT754.',
    )
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    # Round 1 starts at V: 1♣ - P - 2♣ - P
    assert str(ws.cell(row=21, column=2).value) == '1♣'
    assert ws.cell(row=21, column=3).value == 'PAS'
    assert str(ws.cell(row=21, column=4).value) == '2♣'
    assert ws.cell(row=22, column=1).value == 'PAS'

    # Round 2 opener rebid should now show hearts, not 3♣.
    assert str(ws.cell(row=22, column=2).value) == '2♥'


def test_right_info_block_contract_fields():
    """Info block: Kontrakt at C2, Spilfører at C3, Udspil at C4, Resultat at C5."""
    df = _make_df(contract='4S', decl='N', lead='♥A', tricks=10)
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert '4S' in str(ws.cell(row=2, column=3).value)
    assert 'N'  in str(ws.cell(row=3, column=3).value)
    assert '♥A' in str(ws.cell(row=4, column=3).value)
    assert '10'  in str(ws.cell(row=5, column=3).value)


def test_right_info_block_hcp_totals_from_columns():
    """NS/ØV HCP totals shown at A4 and A5 from combined columns when present."""
    df = _make_df(NS_HCP=26, ØV_HCP=14)
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert '26' in str(ws.cell(row=4, column=1).value)
    assert '14' in str(ws.cell(row=5, column=1).value)


def test_right_info_block_hcp_totals_computed():
    """NS/ØV HCP totals at A4 computed from per-hand HCP when combined columns absent."""
    # N_hand=17 HCP, S_hand → compute from 'S_hand'
    # N_hand='AKT7.QJ3.984.AK2' = 17 HCP, S_hand='652.A75.KQ72.J54' = 10 HCP → NS=27
    df = _make_df()
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    ns_cell = str(ws.cell(row=4, column=1).value)
    assert 'NS HCP' in ns_cell
    assert '27' in ns_cell


def test_red_suit_symbols_rich_text():
    """♥ and ♦ suit lines use rich text; ♠ and ♣ remain plain strings."""
    try:
        from openpyxl.cell.rich_text import CellRichText
    except ImportError:
        pytest.skip("openpyxl rich text not available")

    df = _make_df()  # Per = N, N_hand = 'AKT7.QJ3.984.AK2'
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    # Rows 15-18 are the bottom (Per's) hand
    spade_val   = ws.cell(row=15, column=2).value
    heart_val   = ws.cell(row=16, column=2).value
    diamond_val = ws.cell(row=17, column=2).value
    club_val    = ws.cell(row=18, column=2).value

    assert not isinstance(spade_val, CellRichText), "♠ should be plain string"
    assert not isinstance(club_val,  CellRichText), "♣ should be plain string"
    assert isinstance(heart_val,   CellRichText), "♥ should be CellRichText"
    assert isinstance(diamond_val, CellRichText), "♦ should be CellRichText"


def test_spilfoerer_label_at_c3():
    """Label at C3 must say 'Spilfører' (not 'Declarer')."""
    df = _make_df(decl='N')
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    label = str(ws.cell(row=3, column=3).value)
    assert 'Spilfører' in label
    assert 'Declarer' not in label


def test_red_suit_in_info_cells():
    """♥ in lead cell (C4) and contract cell (C2) must be red rich text."""
    try:
        from openpyxl.cell.rich_text import CellRichText
    except ImportError:
        pytest.skip("openpyxl rich text not available")

    df = _make_df(contract='4♥', lead='♥A')
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    contract_cell = ws.cell(row=2, column=3).value  # C2 Kontrakt
    lead_cell = ws.cell(row=4, column=3).value       # C4 Udspil
    assert isinstance(contract_cell, CellRichText), "Contract cell with ♥ should be CellRichText"
    assert isinstance(lead_cell, CellRichText), "Lead cell with ♥ should be CellRichText"


def test_right_info_block_fallback_when_missing():
    """Info block shows '(ukendt)' at C2 and C3 when contract fields absent."""
    df = _make_df()  # no contract/decl/lead/tricks columns
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert '(ukendt)' in str(ws.cell(row=2, column=3).value)
    assert '(ukendt)' in str(ws.cell(row=3, column=3).value)


def test_par_note_moved_to_o9():
    """Par/DD best-contract note should be written in O9 instead of E14."""
    df = _make_df(par_score=-420, par_contract='4♠', par_side='ØV')
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert 'Par:' in str(ws.cell(row=9, column=15).value)
    assert ws.cell(row=14, column=5).value is None


# ---------------------------------------------------------------------------
# Tests for mini traveller table (E1:S2)
# ---------------------------------------------------------------------------

def test_mini_traveller_headers_at_e1():
    """Mini traveller header row must be written at E1:S1."""
    df = _make_df()
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    # Column E = 5; headers span E1..S1 (15 columns)
    headers = [ws.cell(row=1, column=5 + i).value for i in range(15)]
    assert headers[0] == 'NS'
    assert headers[1] == 'ØV'
    assert headers[2] == 'Kontrakt'
    assert headers[3] == 'Spilfører'
    assert headers[4] == 'Udspil'
    assert headers[5] == 'Stik'
    assert headers[6] == 'Score NS'
    assert headers[7] == 'Score ØV'
    assert headers[8] == 'Point NS'
    assert headers[9] == 'Point ØV'
    assert headers[10] == 'Pct NS'
    assert headers[11] == 'Pct ØV'
    assert headers[12] == 'Pct Defense'
    assert headers[13] == 'Pct Decl'
    assert headers[14] == 'Lead type'


def test_mini_traveller_data_row_player_names():
    """Mini traveller data row (row 2) shows combined NS and ØV names."""
    df = _make_df(contract='4S', lead='♥A', tricks=10, pct_NS=58.3, pct_ØV=41.7, decl='N')
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    ns_cell = ws.cell(row=2, column=5).value   # E2 = NS
    ov_cell = ws.cell(row=2, column=6).value   # F2 = ØV
    contract_cell = ws.cell(row=2, column=7).value  # G2 = Kontrakt
    decl_cell = ws.cell(row=2, column=8).value      # H2 = Spilfører
    lead_cell = ws.cell(row=2, column=9).value      # I2 = Udspil
    tricks_cell = ws.cell(row=2, column=10).value   # J2 = Stik
    pct_ns_cell = ws.cell(row=2, column=15).value   # O2 = Pct NS
    pct_def_cell = ws.cell(row=2, column=17).value  # Q2 = Pct Defense
    pct_decl_cell = ws.cell(row=2, column=18).value # R2 = Pct Decl

    assert PER in str(ns_cell)
    assert HENRIK in str(ns_cell)
    assert 'Opp East' in str(ov_cell)
    assert 'Opp West' in str(ov_cell)
    assert contract_cell == '4S'
    assert decl_cell == 'N'
    assert str(lead_cell) == '♥A'
    assert tricks_cell == 10
    assert pct_ns_cell == 58.3
    assert pct_def_cell == 41.7
    assert pct_decl_cell == 58.3


def test_mini_traveller_pct_ov_column():
    """Pct ØV is read from pct_ØV column and placed at P2 (column 16)."""
    df = _make_df(pct_ØV=41.7)
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert ws.cell(row=2, column=16).value == 41.7


def test_mini_traveller_missing_pct_is_none():
    """When pct columns are absent, the data cells are None."""
    df = _make_df()  # no pct columns
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert ws.cell(row=2, column=15).value is None   # Pct NS
    assert ws.cell(row=2, column=16).value is None   # Pct ØV
    assert ws.cell(row=2, column=17).value == 'ukendt'   # Pct Defense
    assert ws.cell(row=2, column=18).value == 'ukendt'   # Pct Decl


# ---------------------------------------------------------------------------
# Tests for Double Dummy table (E5:K9)
# ---------------------------------------------------------------------------

def _make_df_with_dd(**extra):
    """Create a DataFrame row with dd_valid=True and representative DD values."""
    dd_data = {
        'dd_valid': True,
        'dd_N_NT': 9, 'dd_N_S': 10, 'dd_N_H': 8, 'dd_N_D': 7, 'dd_N_C': 6, 'dd_N_HCP': 17,
        'dd_S_NT': 8, 'dd_S_S': 9,  'dd_S_H': 7, 'dd_S_D': 6, 'dd_S_C': 5, 'dd_S_HCP': 10,
        'dd_Ø_NT': 4, 'dd_Ø_S': 3,  'dd_Ø_H': 5, 'dd_Ø_D': 6, 'dd_Ø_C': 7, 'dd_Ø_HCP': 6,
        'dd_V_NT': 5, 'dd_V_S': 4,  'dd_V_H': 6, 'dd_V_D': 7, 'dd_V_C': 8, 'dd_V_HCP': 7,
    }
    dd_data.update(extra)
    return _make_df(**dd_data)


def test_dd_table_title_when_valid():
    """Title cell E5 must read 'Double Dummy' when dd_valid is True."""
    df = _make_df_with_dd()
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert ws.cell(row=5, column=5).value == 'Double Dummy'


def test_dd_table_title_cell_is_gray_header_style():
    """Double Dummy title cell should use same gray fill as DD header cells."""
    df = _make_df_with_dd()
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    dd_title_row = next(
        r for r in range(1, 120)
        if ws.cell(row=r, column=5).value == 'Double Dummy'
    )
    rgb = str(ws.cell(row=dd_title_row, column=5).fill.fgColor.rgb or '')
    assert rgb.endswith('D9D9D9')


def test_dd_table_strain_headers():
    """Strain headers in F5:K5 must be NT, ♠, ♥, ♦, ♣, HP."""
    df = _make_df_with_dd()
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    headers = [str(ws.cell(row=5, column=6 + i).value) for i in range(6)]
    assert headers == ['NT', '♠', '♥', '♦', '♣', 'HP']


def test_dd_table_row_labels():
    """Row labels in E6:E9 must be N, S, Ø, V."""
    df = _make_df_with_dd()
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    labels = [ws.cell(row=6 + i, column=5).value for i in range(4)]
    assert labels == ['N', 'S', 'Ø', 'V']


def test_dd_table_values_populated():
    """DD values must be written correctly in F6:J9."""
    df = _make_df_with_dd()
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    # Row 6 = N:  NT=9, ♠=10, ♥=8, ♦=7, ♣=6
    assert ws.cell(row=6, column=6).value == 9    # dd_N_NT
    assert ws.cell(row=6, column=7).value == 10   # dd_N_S (♠)
    assert ws.cell(row=6, column=8).value == 8    # dd_N_H (♥)
    assert ws.cell(row=6, column=9).value == 7    # dd_N_D (♦)
    assert ws.cell(row=6, column=10).value == 6   # dd_N_C (♣)

    # Row 8 = Ø:  NT=4
    assert ws.cell(row=8, column=6).value == 4    # dd_Ø_NT


def test_dd_table_highlights_target_pair_rows_yellow():
    """DD rows for the target pair directions should be highlighted in yellow."""
    df = _make_df_with_dd(ns1=HENRIK, ns2=PER)
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    dd_title_row = next(
        r for r in range(1, 120)
        if ws.cell(row=r, column=5).value == 'Double Dummy'
    )

    # DD data rows are title+1..title+4 (N,S,Ø,V)
    row_by_dir = {
        str(ws.cell(row=dd_title_row + i, column=5).value): dd_title_row + i
        for i in range(1, 5)
    }

    n_row = row_by_dir['N']
    s_row = row_by_dir['S']
    n_rgb = str(ws.cell(row=n_row, column=6).fill.fgColor.rgb or '')
    s_rgb = str(ws.cell(row=s_row, column=6).fill.fgColor.rgb or '')
    assert n_rgb.endswith('FFF2CC')
    assert s_rgb.endswith('FFF2CC')


def test_dd_table_target_pair_row_labels_stay_gray():
    """DD row-label cells (E-col) should remain gray even when N/S data rows are yellow."""
    df = _make_df_with_dd(ns1=HENRIK, ns2=PER)
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    dd_title_row = next(
        r for r in range(1, 120)
        if ws.cell(row=r, column=5).value == 'Double Dummy'
    )

    # Row labels in column E should keep header-gray fill for all directions.
    for i in range(1, 5):
        rr = dd_title_row + i
        rgb = str(ws.cell(row=rr, column=5).fill.fgColor.rgb or '')
        assert rgb.endswith('D9D9D9')


def test_dd_table_hcp_column():
    """HP column (K = column 11) must contain HCP per direction."""
    df = _make_df_with_dd()
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    assert ws.cell(row=6, column=11).value == 17   # dd_N_HCP
    assert ws.cell(row=7, column=11).value == 10   # dd_S_HCP
    assert ws.cell(row=8, column=11).value == 6    # dd_Ø_HCP
    assert ws.cell(row=9, column=11).value == 7    # dd_V_HCP


def test_dd_table_not_available_when_dd_valid_false():
    """When dd_valid is False, E5 shows 'ikke tilgængelig' message."""
    df = _make_df(dd_valid=False)
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    msg = str(ws.cell(row=5, column=5).value)
    assert 'ikke tilgængelig' in msg


def test_dd_table_not_available_when_dd_valid_missing():
    """When dd_valid column is absent, E5 shows 'ikke tilgængelig' message."""
    df = _make_df()  # no dd_valid column
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    msg = str(ws.cell(row=5, column=5).value)
    assert 'ikke tilgængelig' in msg


# ---------------------------------------------------------------------------
# Tests for fallback tournament selection (Board1_LastTournament)
# ---------------------------------------------------------------------------

def _make_two_date_df(late_date='2026-02-01', early_date='2026-01-15',
                      row_section='A', late_ns2='Someone Else'):
    """Create a two-date DataFrame for fallback testing.

    *late_date* has Per but NOT Henrik (ns2 = late_ns2).
    *early_date* has both Per (ns1) and Henrik (ns2).
    Both rows use the same row/section and board_no=1.
    """
    late_row = {
        'tournament_date': late_date,
        'board_no': 1,
        'row': row_section,
        'ns1': PER,
        'ns2': late_ns2,
        'ew1': 'Opp East',
        'ew2': 'Opp West',
        'N_hand': 'AKT7.QJ3.984.AK2',
        'S_hand': '652.A75.KQ72.J54',
        'Ø_hand': 'QJ93.T862.AT.876',
        'V_hand': '84.K94.J653.QT93',
    }
    early_row = {
        'tournament_date': early_date,
        'board_no': 1,
        'row': row_section,
        'ns1': PER,
        'ns2': HENRIK,
        'ew1': 'Opp East',
        'ew2': 'Opp West',
        'N_hand': 'AKT7.QJ3.984.AK2',
        'S_hand': '652.A75.KQ72.J54',
        'Ø_hand': 'QJ93.T862.AT.876',
        'V_hand': '84.K94.J653.QT93',
    }
    return pd.DataFrame([late_row, early_row])


def test_fallback_uses_earlier_date_traveller():
    """When latest tournament lacks Henrik, traveller uses the earlier date's rows."""
    df = _make_two_date_df()
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    # Traveller data row is at row 2, column 5 (NS-par / E2).
    # The early date row has both Per and Henrik as NS.
    ns_cell = str(ws.cell(row=2, column=5).value)
    assert HENRIK in ns_cell, "Traveller should show earlier date row with Henrik"


def test_fallback_note_written_to_sheet():
    """When fallback is applied, a note is written to row 7, column B."""
    df = _make_two_date_df(late_date='2026-02-01', early_date='2026-01-15')
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    note = ws.cell(row=7, column=2).value
    assert note is not None, "Fallback note should be present at row 7, col B"
    note_str = str(note)
    assert '2026-02-01' in note_str, "Note should mention the latest (skipped) date"
    assert '2026-01-15' in note_str, "Note should mention the fallback date"


def test_no_fallback_note_when_both_present():
    """When both names are in the latest tournament, no fallback note is written."""
    # Latest date has both Per (ns1) and Henrik (ns2)
    df = _make_df()  # default: ns1=PER, ns2=HENRIK, date='2026-01-15'
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    note = ws.cell(row=7, column=2).value
    assert note is None, "No fallback note should appear when both players are present"


def test_fallback_no_earlier_date_keeps_original():
    """When no earlier date has both players, the original latest date is used (no crash)."""
    # Only one date in df, and Henrik is missing → no fallback available
    df = _make_df(ns1=PER, ns2='Someone Else', ew1='Opp East', ew2='Opp West')
    writer, wb = _make_writer_mock()
    write_board1_layout_sheet(writer, df, PER)
    ws = wb['Board1_LastTournament']

    # Sheet should be created without error
    assert 'Board1_LastTournament' in wb.sheetnames
    # No fallback note because no earlier date exists
    note = ws.cell(row=7, column=2).value
    assert note is None, "No note when no earlier date with both players exists"


# ---------------------------------------------------------------------------
# Unit tests for _both_names_in_df helper
# ---------------------------------------------------------------------------

def test_both_names_found():
    df = _make_df()  # ns1=PER, ns2=HENRIK
    assert _both_names_in_df(df, PER, HENRIK) is True


def test_both_names_one_missing():
    df = _make_df(ns2='Someone Else')
    assert _both_names_in_df(df, PER, HENRIK) is False


def test_both_names_strips_whitespace():
    df = _make_df(ns1=' Per Føge Jensen ', ns2=' Henrik Friis ')
    assert _both_names_in_df(df, PER, HENRIK) is True


def test_both_names_handles_none():
    df = _make_df(ns1=None, ns2=None, ew1=PER, ew2=HENRIK)
    assert _both_names_in_df(df, PER, HENRIK) is True

